"""xls_create_from_template: Clone from .xltx/.xltm template with variable substitution."""

from __future__ import annotations

import re

from openpyxl import load_workbook

from excel_agent.core.version_hash import compute_file_hash
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    create_parser,
    parse_json_arg,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response

# Regex for {{placeholder}} patterns
_PLACEHOLDER_RE = re.compile(r"\{\{(\w+)\}\}")


def _run() -> dict:
    parser = create_parser(
        "Create a new workbook from a template (.xltx/.xltm) with variable substitution."
    )
    parser.add_argument(
        "--template", type=str, required=True, help="Path to template file (.xltx, .xltm, .xlsx)"
    )
    parser.add_argument("--output", type=str, required=True, help="Output workbook path (.xlsx)")
    parser.add_argument(
        "--vars",
        type=str,
        default="{}",
        help='JSON object of variable substitutions (e.g., \'{"company": "Acme"}\')',
    )
    args = parser.parse_args()

    template_path = validate_input_path(args.template)
    output_path = validate_output_path(args.output, create_parents=True)
    variables: dict = parse_json_arg(args.vars)

    # Detect if template has macros (.xltm or .xlsm)
    template_ext = template_path.suffix.lower()
    keep_vba = template_ext in {".xltm", ".xlsm"}

    # Load template with macro preservation if applicable
    wb = load_workbook(str(template_path), keep_vba=keep_vba)
    wb.template = False

    warnings = []

    # Check for macro loss when converting .xltm/.xlsm to .xlsx
    output_ext = output_path.suffix.lower()
    if keep_vba and output_ext not in {".xlsm", ".xltm"}:
        warnings.append(
            f"Template contains macros but output is {output_ext}. "
            f"Use .xlsm extension to preserve macros."
        )

    # Perform variable substitution across all sheets
    substitutions_made = 0
    substitution_details: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f" and isinstance(cell.value, str):
                    # Only substitute in text cells, NOT in formulas
                    original = cell.value
                    new_value = _substitute_placeholders(original, variables)
                    if new_value != original:
                        cell.value = new_value
                        substitutions_made += 1
                        substitution_details.append(
                            {
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "original": original,
                                "replaced": new_value,
                            }
                        )

    wb.save(str(output_path))
    file_hash = compute_file_hash(output_path)

    return build_response(
        "success",
        {
            "output_path": str(output_path),
            "template_path": str(template_path),
            "variables": variables,
            "substitutions_made": substitutions_made,
            "substitution_details": substitution_details[:20],
            "sheets": list(wb.sheetnames),
        },
        workbook_version=file_hash,
        impact={"cells_modified": substitutions_made, "formulas_updated": 0},
        warnings=warnings if warnings else None,
    )


def _substitute_placeholders(text: str, variables: dict) -> str:
    """Replace {{key}} placeholders with values from the variables dict.

    Unmatched placeholders are left as-is (not replaced).
    """

    def _replacer(match: re.Match) -> str:  # type: ignore[type-arg]
        key = match.group(1)
        if key in variables:
            return str(variables[key])
        return match.group(0)  # Leave unmatched placeholders as-is

    return _PLACEHOLDER_RE.sub(_replacer, text)


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()

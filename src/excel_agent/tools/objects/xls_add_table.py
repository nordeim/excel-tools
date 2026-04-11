"""xls_add_table: Convert range to Excel Table (ListObject).

Creates an Excel Table from a data range with styling and structured references.
Tables provide auto-filtering, sorting, and structured reference capabilities.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_agent.core.edit_session import EditSession
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    check_macro_contract,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response

# Built-in table styles (60 total)
TABLE_STYLES = [
    # Light styles (21)
    *[f"TableStyleLight{i}" for i in range(1, 22)],
    # Medium styles (28)
    *[f"TableStyleMedium{i}" for i in range(1, 29)],
    # Dark styles (11)
    *[f"TableStyleDark{i}" for i in range(1, 12)],
]

DEFAULT_STYLE = "TableStyleMedium9"


def _is_valid_table_name(name: str) -> bool:
    """Check if table name is valid for Excel.

    Rules:
    - Must start with letter or underscore
    - Can contain letters, numbers, underscores
    - No spaces
    - Max 255 chars
    """
    if not name:
        return False
    if len(name) > 255:
        return False
    if " " in name:
        return False
    # Must start with letter or underscore
    if not re.match(r"^[a-zA-Z_][a-zA-Z0-9_]*$", name):
        return False
    return True


def _get_table_columns(ws, min_col: int, max_col: int, header_row: int) -> list[str]:
    """Extract column names from header row."""
    columns = []
    for col_idx in range(min_col, max_col + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value:
            columns.append(str(cell_value))
        else:
            # Generate column name from column letter
            from openpyxl.utils import get_column_letter

            columns.append(f"Column_{get_column_letter(col_idx)}")
    return columns


def _run() -> dict[str, object]:
    parser = create_parser("Convert range to Excel Table with styling.")
    add_common_args(parser)
    parser.add_argument(
        "--range",
        type=str,
        required=True,
        help='Table range (e.g., "A1:D100") - must include headers',
    )
    parser.add_argument(
        "--name",
        type=str,
        required=True,
        help="Table display name (must be unique, no spaces)",
    )
    parser.add_argument(
        "--style",
        type=str,
        default=DEFAULT_STYLE,
        help=f"Table style (default: {DEFAULT_STYLE})",
    )
    parser.add_argument(
        "--show-headers",
        action="store_true",
        default=True,
        help="Show table headers (default: True)",
    )
    parser.add_argument(
        "--show-row-stripes",
        action="store_true",
        default=True,
        help="Show alternating row colors (default: True)",
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(
        args.output or str(input_path),
        create_parents=True,
    )

    # Check for macro loss warning
    macro_warning = check_macro_contract(input_path, output_path)
    warnings = [macro_warning] if macro_warning else []

    # Validate table name
    if not _is_valid_table_name(args.name):
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[
                "Invalid table name. Must:",
                "- Start with letter or underscore",
                "- Contain only letters, numbers, underscores",
                "- Have no spaces",
                "- Be max 255 characters",
            ],
        )

    # Validate style
    if args.style not in TABLE_STYLES:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[
                f"Invalid table style: {args.style}",
                f"Valid styles: {', '.join(TABLE_STYLES[:5])}... (60 total)",
            ],
        )

    # Use EditSession for proper locking and save semantics
    session = EditSession.prepare(input_path, output_path)

    with session:
        wb = session.workbook
        ws = wb[args.sheet] if args.sheet else wb.active

        # Check if table name already exists
        for table in ws.tables.values():
            if table.displayName == args.name:
                return build_response(
                    "error",
                    None,
                    exit_code=1,
                    warnings=[
                        f"Table name '{args.name}' already exists in this workbook",
                        "Table names must be unique workbook-wide",
                    ],
                )

        # Parse range
        try:
            from openpyxl.utils import range_boundaries

            min_col, min_row, max_col, max_row = range_boundaries(args.range)
            if min_col is None or min_row is None:
                raise ValueError("Invalid range format")
            # Handle single-cell case
            if max_col is None:
                max_col = min_col
            if max_row is None:
                max_row = min_row
        except Exception as e:
            return build_response(
                "error",
                None,
                exit_code=1,
                warnings=[f"Failed to parse range '{args.range}': {e}"],
            )

        # Validate range has at least header row + one data row
        if max_row < min_row + 1:
            warnings.append("Range has only one row. Tables typically need headers + data.")

        # Check for overlapping tables
        for existing_table in ws.tables.values():
            existing_ref = existing_table.ref
            from openpyxl.utils import range_boundaries

            ec1, er1, ec2, er2 = range_boundaries(existing_ref)
            # Check overlap
            if not (max_col < ec1 or min_col > ec2 or max_row < er1 or min_row > er2):
                return build_response(
                    "error",
                    None,
                    exit_code=1,
                    warnings=[
                        f"Range {args.range} overlaps with existing table "
                        f"'{existing_table.displayName}'",
                        "Table ranges must not overlap",
                    ],
                )

        # Create table
        table = Table(displayName=args.name, ref=args.range)

        # Apply style
        style = TableStyleInfo(
            name=args.style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=args.show_row_stripes,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style

        # Add table to worksheet
        ws.add_table(table)

        # Capture version hash before exiting context
        version_hash = session.version_hash

        # EditSession handles save automatically on exit

    # Log to audit trail (after successful save)
    audit = AuditTrail()
    audit.log(
        tool="xls_add_table",
        scope="structure:modify",
        target_file=input_path,
        file_version_hash=session.file_hash,
        actor_nonce="auto",
        operation_details={
            "table_name": args.name,
            "range": args.range,
            "style": args.style,
            "sheet": ws.title,
        },
        impact={
            "table_created": True,
            "row_count": max_row - min_row + 1,
            "column_count": max_col - min_col + 1,
        },
        success=True,
        exit_code=0,
    )

    return build_response(
        "success",
        {
            "table_name": args.name,
            "range": args.range,
            "sheet": ws.title,
            "style": args.style,
            "row_count": max_row - min_row + 1,
            "column_count": max_col - min_col + 1,
            "columns": _get_table_columns(ws, min_col, max_col, min_row),
        },
        workbook_version=version_hash,
        warnings=warnings if warnings else None,
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()

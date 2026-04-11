"""xls_delete_columns: Delete columns with dependency check (token required)."""

from __future__ import annotations

from openpyxl.utils import column_index_from_string, get_column_letter

from excel_agent.core.dependency import DependencyTracker
from excel_agent.core.edit_session import EditSession
from excel_agent.core.formula_updater import adjust_col_references
from excel_agent.core.version_hash import compute_file_hash
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.governance.token_manager import ApprovalTokenManager
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    add_governance_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.exceptions import ImpactDeniedError, ValidationError
from excel_agent.utils.json_io import build_response


def _parse_column(col_str: str) -> int:
    try:
        return int(col_str)
    except ValueError:
        return column_index_from_string(col_str.upper())


def _run() -> dict:
    parser = create_parser(
        "Delete columns from a worksheet. Requires an approval token (scope: range:delete)."
    )
    add_common_args(parser)
    add_governance_args(parser)
    parser.add_argument(
        "--start-column", type=str, required=True, help="First column (letter or number)"
    )
    parser.add_argument("--count", type=int, default=1, help="Number of columns (default: 1)")
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(args.output or args.input, create_parents=True)
    file_hash = compute_file_hash(input_path)

    if not args.token:
        raise ValidationError("Approval token required for column deletion.")
    mgr = ApprovalTokenManager()
    mgr.validate_token(args.token, expected_scope="range:delete", expected_file_hash=file_hash)

    start_col = _parse_column(args.start_column)

    # Use EditSession for proper copy-on-write and save semantics
    session = EditSession.prepare(input_path, output_path)

    with session:
        wb = session.workbook
        sheet_name = args.sheet or wb.sheetnames[0]
        ws = wb[sheet_name]
        end_col = start_col + args.count - 1
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)

        tracker = DependencyTracker(wb)
        tracker.build_graph()
        target = f"{sheet_name}!{start_letter}1:{end_letter}1048576"
        report = tracker.impact_report(target, action="delete")

        if report.broken_references > 0 and not args.acknowledge_impact:
            raise ImpactDeniedError(
                f"Deleting columns {start_letter}-{end_letter} would break "
                f"{report.broken_references} formula reference(s)",
                impact_report=report.to_dict(),
                guidance="Run xls-update-references first, or use --acknowledge-impact.",
            )

        ws.delete_cols(idx=start_col, amount=args.count)
        formulas_updated = adjust_col_references(wb, sheet_name, start_col, -args.count)

        # Capture version hash before exiting context
        version_hash = session.version_hash

        # EditSession handles save automatically on exit

    audit = AuditTrail()
    audit.log_operation(
        tool="xls_delete_columns",
        scope="range:delete",
        resource=f"{sheet_name}!cols {start_letter}-{end_letter}",
        action="delete",
        outcome="success",
        token_used=True,
        file_hash=session.file_hash,
    )

    return build_response(
        "success",
        {
            "sheet": sheet_name,
            "start_column": args.start_column,
            "columns_deleted": args.count,
            "impact": report.to_dict(),
        },
        workbook_version=version_hash,
        impact={"cells_modified": 0, "formulas_updated": formulas_updated},
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()

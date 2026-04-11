"""xls_format_range: Apply fonts, fills, borders, and alignment to cell ranges.

Applies comprehensive cell styling from JSON specifications including
fonts, fills, borders, and alignment settings.
"""

from __future__ import annotations

import json

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import range_boundaries

from excel_agent.core.edit_session import EditSession
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    check_macro_contract,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response

# Maximum cells before performance warning
MAX_CELLS_WARNING = 10000


def _parse_json_style(style_str: str | None) -> dict:
    """Parse JSON style string to dict."""
    if not style_str:
        return {}

    try:
        return json.loads(style_str)
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON: {e}")


def _create_font(font_spec: dict) -> Font:
    """Create Font object from specification."""
    return Font(
        name=font_spec.get("name", "Calibri"),
        size=font_spec.get("size", 11),
        bold=font_spec.get("bold", False),
        italic=font_spec.get("italic", False),
        underline=font_spec.get("underline"),
        color=font_spec.get("color"),
    )


def _create_fill(fill_spec: dict) -> PatternFill:
    """Create PatternFill object from specification."""
    pattern_type = fill_spec.get("patternType", "solid")
    if pattern_type == "solid":
        return PatternFill(
            start_color=fill_spec.get("fgColor", "FFFFFF"),
            end_color=fill_spec.get("bgColor", fill_spec.get("fgColor", "FFFFFF")),
            fill_type="solid",
        )
    return PatternFill(
        patternType=pattern_type,
        fgColor=fill_spec.get("fgColor"),
        bgColor=fill_spec.get("bgColor"),
    )


def _create_side(side_spec: dict | None) -> Side | None:
    """Create Side object from specification."""
    if not side_spec:
        return None
    return Side(style=side_spec.get("style", "thin"), color=side_spec.get("color", "000000"))


def _create_border(border_spec: dict) -> Border:
    """Create Border object from specification."""
    return Border(
        left=_create_side(border_spec.get("left")),
        right=_create_side(border_spec.get("right")),
        top=_create_side(border_spec.get("top")),
        bottom=_create_side(border_spec.get("bottom")),
    )


def _create_alignment(align_spec: dict) -> Alignment:
    """Create Alignment object from specification."""
    return Alignment(
        horizontal=align_spec.get("horizontal", "left"),
        vertical=align_spec.get("vertical", "bottom"),
        wrap_text=align_spec.get("wrapText", False),
        text_rotation=align_spec.get("textRotation", 0),
    )


def _run() -> dict[str, object]:
    parser = create_parser("Apply fonts, fills, borders, and alignment to cell ranges.")
    add_common_args(parser)
    parser.add_argument(
        "--range",
        type=str,
        required=True,
        help='Target range (e.g., "A1:D10")',
    )
    parser.add_argument(
        "--font",
        type=str,
        default=None,
        help='Font JSON: {"name": "Calibri", "size": 11, "bold": true, "color": "FF0000"}',
    )
    parser.add_argument(
        "--fill",
        type=str,
        default=None,
        help='Fill JSON: {"fgColor": "FFFF00", "patternType": "solid"}',
    )
    parser.add_argument(
        "--border",
        type=str,
        default=None,
        help='Border JSON: {"top": {"style": "thin", "color": "000000"}}',
    )
    parser.add_argument(
        "--alignment",
        type=str,
        default=None,
        help='Alignment JSON: {"horizontal": "center", "vertical": "middle", "wrapText": true}',
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(
        args.output or str(input_path),
        create_parents=True,
    )

    # Check for macro loss warning
    macro_warning = check_macro_contract(input_path, output_path)
    warnings = [macro_warning] if macro_warning else []

    # Parse style specifications
    try:
        font_spec = _parse_json_style(args.font)
        fill_spec = _parse_json_style(args.fill)
        border_spec = _parse_json_style(args.border)
        alignment_spec = _parse_json_style(args.alignment)
    except ValueError as e:
        return build_response("error", None, exit_code=1, warnings=[str(e)])

    # Validate at least one style is provided
    if not any([font_spec, fill_spec, border_spec, alignment_spec]):
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=["At least one style (font, fill, border, alignment) must be provided"],
        )

    # Use EditSession for proper locking and save semantics
    session = EditSession.prepare(input_path, output_path)

    with session:
        wb = session.workbook
        ws = wb[args.sheet] if args.sheet else wb.active
        if ws is None:
            return build_response("error", None, exit_code=1, warnings=["No active sheet found"])

        # Parse range
        try:
            min_col, min_row, max_col, max_row = range_boundaries(args.range)
            if min_col is None or min_row is None:
                raise ValueError("Invalid range format")
            if max_col is None:
                max_col = min_col
            if max_row is None:
                max_row = min_row
        except Exception as e:
            return build_response(
                "error",
                None,
                exit_code=1,
                warnings=[f"Failed to parse range '{args.range}': {e}"],
            )

        # Check cell count for performance warning
        cell_count = (max_col - min_col + 1) * (max_row - min_row + 1)
        if cell_count > MAX_CELLS_WARNING:
            warnings.append(
                f"Formatting {cell_count} cells may be slow. Consider processing in batches."
            )

        # Create style objects
        font = _create_font(font_spec) if font_spec else None
        fill = _create_fill(fill_spec) if fill_spec else None
        border = _create_border(border_spec) if border_spec else None
        alignment = _create_alignment(alignment_spec) if alignment_spec else None

        # Apply styles to range
        cells_formatted = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if border:
                    cell.border = border
                if alignment:
                    cell.alignment = alignment
                cells_formatted += 1

        # Capture version hash before exiting context
        version_hash = session.version_hash

        # EditSession handles save automatically on exit

    # Log to audit trail (after successful save)
    audit = AuditTrail()
    audit.log(
        tool="xls_format_range",
        scope="structure:modify",
        target_file=input_path,
        file_version_hash=session.file_hash,
        actor_nonce="auto",
        operation_details={
            "range": args.range,
            "sheet": ws.title,
            "font_applied": bool(font),
            "fill_applied": bool(fill),
            "border_applied": bool(border),
            "alignment_applied": bool(alignment),
        },
        impact={"cells_formatted": cells_formatted},
        success=True,
        exit_code=0,
    )

    return build_response(
        "success",
        {
            "range": args.range,
            "sheet": ws.title,
            "cells_formatted": cells_formatted,
            "font_applied": bool(font),
            "fill_applied": bool(fill),
            "border_applied": bool(border),
            "alignment_applied": bool(alignment),
        },
        workbook_version=version_hash,
        warnings=warnings if warnings else None,
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()

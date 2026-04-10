"""xls_set_number_format: Apply number formats to cell ranges.

Supports Excel format codes including:
- Currency: '"$"#,##0.00'
- Percentage: '0.00%'
- Date: 'yyyy-mm-dd', 'mm/dd/yyyy'
- Time: 'h:mm:ss'
- Scientific: '0.00E+00'
- Fraction: '# ?/?'
- Custom: Any valid Excel format code

Format codes are passed directly to openpyxl (no abstraction layer).
"""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from excel_agent.core.version_hash import compute_file_hash
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response

# Common format examples for documentation
COMMON_FORMATS = {
    '"$"#,##0.00': "Currency with dollar sign, thousands separator, 2 decimals",
    '"€"#,##0.00': "Euro currency format",
    '"£"#,##0.00': "Pound currency format",
    "0.00%": "Percentage with 2 decimals",
    "0.0%": "Percentage with 1 decimal",
    "yyyy-mm-dd": "Date (ISO format)",
    "mm/dd/yyyy": "Date (US format)",
    "dd/mm/yyyy": "Date (European format)",
    "h:mm:ss": "Time (hours, minutes, seconds)",
    "h:mm": "Time (hours, minutes)",
    "0.00E+00": "Scientific notation",
    "#,##0": "Number with thousands separator",
    "#,##0.00": "Number with thousands separator, 2 decimals",
    "# ?/?": "Fraction (up to 1 digit)",
    "# ??/??": "Fraction (up to 2 digits)",
    "0": "Whole number",
    "0.0": "Number with 1 decimal",
    "@": "Text format",
}


def _run() -> dict[str, object]:
    parser = create_parser("Apply number formats to cell ranges.")
    add_common_args(parser)
    parser.add_argument(
        "--range",
        type=str,
        required=True,
        help='Target range (e.g., "A1:A100")',
    )
    parser.add_argument(
        "--number-format",
        type=str,
        required=True,
        dest="number_format",
        help="Excel number format code (e.g., '\"$\"#,##0.00', '0.00%%', 'yyyy-mm-dd')",
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(args.output or str(input_path), create_parents=True)

    file_hash = compute_file_hash(input_path)

    # Validate format string
    if not args.number_format.strip():
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=["Format string cannot be empty"],
        )

    # Load workbook
    wb = load_workbook(str(input_path))
    ws = wb[args.sheet] if args.sheet else wb.active
    if ws is None:
        return build_response("error", None, exit_code=1, warnings=["No active sheet found"])

    # Parse range
    try:
        min_col, min_row, max_col, max_row = range_boundaries(args.range)
        if min_col is None or min_row is None:
            raise ValueError("Invalid range format")
        if max_col is None:
            max_col = min_col
        if max_row is None:
            max_row = min_row
    except Exception as e:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[f"Failed to parse range '{args.range}': {e}"],
        )

    # Apply format to range
    cells_formatted = 0
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = args.number_format
            cells_formatted += 1

    # Save workbook
    wb.save(str(output_path))

    # Log to audit trail
    audit = AuditTrail()
    audit.log(
        tool="xls_set_number_format",
        scope="structure:modify",
        target_file=input_path,
        file_version_hash=file_hash,
        actor_nonce="auto",
        operation_details={
            "range": args.range,
            "format": args.number_format,
            "sheet": ws.title,
            "cells_formatted": cells_formatted,
        },
        impact={"cells_formatted": cells_formatted},
        success=True,
        exit_code=0,
    )

    # Check if format is in common formats for guidance
    format_description = COMMON_FORMATS.get(args.number_format, "Custom format")

    return build_response(
        "success",
        {
            "range": args.range,
            "format": args.number_format,
            "format_description": format_description,
            "sheet": ws.title,
            "cells_formatted": cells_formatted,
        },
        warnings=None,
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()

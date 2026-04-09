"""xls_merge_cells: Merge a range of cells with hidden data pre-check.

When cells are merged, all cells but the top-left one are removed from
the worksheet (openpyxl 3.1.5 docs: "When you merge cells all cells
but the top-left one are removed from the worksheet"). Non-anchor cells
become MergeCells with value=None.

This tool checks for non-empty values in non-anchor cells BEFORE merging
and returns a warning. The agent must use --force to proceed if data
would be lost.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter

from excel_agent.core.agent import ExcelAgent
from excel_agent.core.serializers import RangeSerializer
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    add_governance_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.exceptions import ValidationError
from excel_agent.utils.json_io import build_response


def _run() -> dict[str, object]:
    parser = create_parser("Merge a range of cells. Warns if non-anchor cells contain data.")
    add_common_args(parser)
    add_governance_args(parser)
    parser.add_argument("--range", type=str, required=True, help="Range to merge (e.g., A1:C1)")
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(args.output or args.input, create_parents=True)

    with ExcelAgent(input_path, mode="rw") as agent:
        wb = agent.workbook
        serializer = RangeSerializer(workbook=wb)
        coord = serializer.parse(args.range, default_sheet=args.sheet)

        sheet_name = coord.sheet or (args.sheet or wb.sheetnames[0])
        ws = wb[sheet_name]

        if coord.max_row is None or coord.max_col is None:
            raise ValidationError("Merge requires a range, not a single cell.")

        min_row = coord.min_row
        min_col = coord.min_col
        max_row = coord.max_row
        max_col = coord.max_col

        # Pre-check: scan non-anchor cells for data
        hidden_data: list[dict[str, object]] = []
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                if row_idx == min_row and col_idx == min_col:
                    continue  # Skip anchor cell (top-left)
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    hidden_data.append(
                        {
                            "cell": f"{get_column_letter(col_idx)}{row_idx}",
                            "value": str(cell.value)[:100],
                        }
                    )

        warnings: list[str] = []
        if hidden_data and not args.force:
            return build_response(
                "warning",
                {
                    "merge_range": args.range,
                    "hidden_data": hidden_data[:10],
                    "data_loss_count": len(hidden_data),
                },
                exit_code=1,
                warnings=[
                    f"{len(hidden_data)} non-anchor cell(s) contain data that will be lost. "
                    f"Use --force to proceed."
                ],
                guidance="Use --force to merge despite data in non-anchor cells.",
                workbook_version=agent.version_hash,
            )

        if hidden_data:
            warnings.append(
                f"{len(hidden_data)} non-anchor cell(s) had data that was discarded "
                "(--force used)."
            )

        # Perform the merge
        range_string = (
            f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        )
        ws.merge_cells(range_string)

        if str(output_path) != str(input_path):
            wb.save(str(output_path))

        return build_response(
            "success",
            {
                "merge_range": range_string,
                "sheet": sheet_name,
                "anchor_cell": f"{get_column_letter(min_col)}{min_row}",
                "data_discarded": len(hidden_data),
            },
            workbook_version=agent.version_hash,
            impact={
                "cells_modified": (max_row - min_row + 1) * (max_col - min_col + 1),
                "formulas_updated": 0,
            },
            warnings=warnings,
        )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()

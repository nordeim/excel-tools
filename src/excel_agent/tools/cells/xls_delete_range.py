"""xls_delete_range: Clear a range and shift cells up or left (token required).

Uses openpyxl's move_range to shift remaining cells into the gap.
Per openpyxl docs: "Move a cell range by the number of rows and/or
columns: down if rows > 0 and up if rows < 0, right if cols > 0 and
left if cols < 0. Existing cells will be overwritten."

With translate=True, formulae in the MOVED cells are translated, but
references FROM OTHER cells are NOT updated — our formula_updater
handles that.

Requires approval token (scope: range:delete) and performs a
pre-flight dependency impact check.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter

from excel_agent.core.agent import ExcelAgent
from excel_agent.core.dependency import DependencyTracker
from excel_agent.core.formula_updater import adjust_col_references, adjust_row_references
from excel_agent.core.serializers import RangeSerializer
from excel_agent.core.version_hash import compute_file_hash
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.governance.token_manager import ApprovalTokenManager
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    add_governance_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.exceptions import ImpactDeniedError, ValidationError
from excel_agent.utils.json_io import build_response


def _run() -> dict[str, Any]:
    parser = create_parser(
        "Delete a range of cells and shift remaining cells up or left. "
        "Requires approval token (scope: range:delete)."
    )
    add_common_args(parser)
    add_governance_args(parser)
    parser.add_argument("--range", type=str, required=True, help="Range to delete (e.g., A5:C10)")
    parser.add_argument(
        "--shift",
        type=str,
        required=True,
        choices=["up", "left"],
        help="Direction to shift remaining cells: 'up' or 'left'",
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(args.output or args.input, create_parents=True)
    file_hash = compute_file_hash(input_path)

    if not args.token:
        raise ValidationError(
            "Approval token required for range deletion. "
            "Generate one with: xls-approve-token --scope range:delete --file <path>"
        )
    mgr = ApprovalTokenManager()
    mgr.validate_token(args.token, "range:delete", input_path)

    with ExcelAgent(input_path, mode="rw") as agent:
        wb = agent.workbook
        serializer = RangeSerializer(workbook=wb)
        coord = serializer.parse(args.range, default_sheet=args.sheet)

        sheet_name = coord.sheet or (args.sheet or wb.sheetnames[0])
        ws = wb[sheet_name]

        min_row = coord.min_row
        min_col = coord.min_col
        max_row = coord.max_row or min_row
        max_col = coord.max_col or min_col

        # Pre-flight dependency check
        tracker = DependencyTracker(wb)
        tracker.build_graph()
        target_str = (
            f"{sheet_name}!{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )
        report = tracker.impact_report(target_str, action="delete")

        if report.broken_references > 0 and not args.acknowledge_impact:
            raise ImpactDeniedError(
                f"Deleting range {args.range} in {sheet_name!r} would break "
                f"{report.broken_references} formula reference(s)",
                impact_report=report.to_dict(),
                guidance=(
                    "Run xls-update-references to fix references first, or re-run with "
                    "--acknowledge-impact and a valid token to proceed."
                ),
            )

        # Clear the cells in the range
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                ws.cell(row=row_idx, column=col_idx).value = None

        # Shift remaining cells
        formulas_updated = 0
        cells_shifted = 0
        row_span = max_row - min_row + 1
        col_span = max_col - min_col + 1

        if args.shift == "up" and ws.max_row and ws.max_row > max_row:
            # Move cells below the deleted range upward
            move_range = (
                f"{get_column_letter(min_col)}{max_row + 1}:"
                f"{get_column_letter(max_col)}{ws.max_row}"
            )
            ws.move_range(move_range, rows=-row_span, translate=True)
            cells_shifted = (ws.max_row - max_row) * col_span
            formulas_updated = adjust_row_references(wb, sheet_name, max_row + 1, -row_span)
        elif args.shift == "left" and ws.max_column and ws.max_column > max_col:
            # Move cells to the right of the deleted range leftward
            move_range = (
                f"{get_column_letter(max_col + 1)}{min_row}:"
                f"{get_column_letter(ws.max_column)}{max_row}"
            )
            ws.move_range(move_range, cols=-col_span, translate=True)
            cells_shifted = (max_row - min_row + 1) * (ws.max_column - max_col)
            formulas_updated = adjust_col_references(wb, sheet_name, max_col + 1, -col_span)

        if str(output_path) != str(input_path):
            wb.save(str(output_path))

        audit = AuditTrail()
        audit.log(
            tool="xls_delete_range",
            scope="range:delete",
            target_file=Path(input_path),
            file_version_hash=file_hash,
            actor_nonce=args.token[:32] if args.token else "unknown",
            operation_details={"shift": args.shift, "cells_shifted": cells_shifted},
            impact={"cells_modified": row_span * col_span + cells_shifted},
            success=True,
            exit_code=0,
        )

        return build_response(
            "success",
            {
                "range": args.range,
                "sheet": sheet_name,
                "shift": args.shift,
                "cells_cleared": row_span * col_span,
                "cells_shifted": cells_shifted,
                "impact": report.to_dict(),
            },
            workbook_version=agent.version_hash,
            impact={
                "cells_modified": row_span * col_span + cells_shifted,
                "formulas_updated": formulas_updated,
            },
        )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()

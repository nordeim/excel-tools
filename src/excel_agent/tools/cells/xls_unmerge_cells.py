"""xls_unmerge_cells: Restore grid from merged range.

Unmerges cells in the specified range. After unmerging, the previously
merged boundary cells (MergeCells) will have value=None — only the
anchor cell retains its value.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter

from excel_agent.core.edit_session import EditSession
from excel_agent.core.serializers import RangeSerializer
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response


def _run() -> dict[str, object]:
    parser = create_parser("Unmerge cells in a range, restoring individual cells.")
    add_common_args(parser)
    parser.add_argument(
        "--range",
        type=str,
        default=None,
        help=("Specific range to unmerge (e.g., A1:C1). If omitted, unmerges ALL merged ranges."),
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(args.output or args.input, create_parents=True)

    session = EditSession.prepare(input_path, output_path)
    with session:
        wb = session.workbook
        sheet_name = args.sheet or wb.sheetnames[0]
        ws = wb[sheet_name]

        unmerged_ranges: list[str] = []

        if args.range:
            # Unmerge a specific range
            serializer = RangeSerializer(workbook=wb)
            coord = serializer.parse(args.range, default_sheet=sheet_name)
            range_string = (
                f"{get_column_letter(coord.min_col)}{coord.min_row}:"
                f"{get_column_letter(coord.max_col or coord.min_col)}"
                f"{coord.max_row or coord.min_row}"
            )
            ws.unmerge_cells(range_string)
            unmerged_ranges.append(range_string)
        else:
            # Unmerge ALL merged ranges on the sheet
            # Collect ranges first to avoid modifying during iteration
            merged_list = list(ws.merged_cells.ranges)
            for merged_range in merged_list:
                range_str = str(merged_range)
                ws.unmerge_cells(range_str)
                unmerged_ranges.append(range_str)

        # Capture version hash before exiting context
        version_hash = session.version_hash

        # EditSession handles save automatically on exit

        return build_response(
            "success",
            {
                "sheet": sheet_name,
                "unmerged_ranges": unmerged_ranges,
                "count": len(unmerged_ranges),
            },
            workbook_version=version_hash,
            impact={"cells_modified": len(unmerged_ranges), "formulas_updated": 0},
        )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()

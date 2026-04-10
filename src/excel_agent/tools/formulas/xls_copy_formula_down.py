"""xls_copy_formula_down: Copy formula down with reference adjustments."""

from __future__ import annotations

from excel_agent.core.agent import ExcelAgent
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response
from openpyxl.utils import get_column_letter


def _run() -> dict:
    parser = create_parser("Copy formula from source cell down to target cells.")
    add_common_args(parser)
    parser.add_argument(
        "--source", type=str, help="Source cell (e.g., A1) - preferred over --cell"
    )
    parser.add_argument("--cell", type=str, help="Source cell (deprecated, use --source)")
    parser.add_argument("--target", type=str, help="Target range (e.g., A1:A10)")
    parser.add_argument(
        "--count", type=int, help="Number of cells to fill (deprecated, use --target)"
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(args.output)

    # Handle both APIs (prefer --source over --cell)
    source = args.source or args.cell
    if not source:
        parser.error("--source (or deprecated --cell) is required")

    # Parse target or count
    if args.target:
        # Parse range to get count
        from openpyxl.utils import range_boundaries

        try:
            min_col, min_row, max_col, max_row = range_boundaries(args.target)
            if min_row and max_row:
                count = max_row - min_row + 1
            else:
                count = 1
        except Exception:
            count = 1
    elif args.count:
        count = args.count
    else:
        parser.error("--target or --count is required")

    formulas_copied = 0

    with ExcelAgent(input_path, mode="rw") as agent:
        wb = agent.workbook
        sheet_name = args.sheet or wb.sheetnames[0]
        ws = wb[sheet_name]

        # Parse source cell
        from excel_agent.core.serializers import RangeSerializer

        serializer = RangeSerializer(wb)
        coord = serializer.parse(source, default_sheet=sheet_name)
        source_row = coord.min_row
        source_col = coord.min_col

        # Get source formula
        source_cell = ws.cell(row=source_row, column=source_col)
        if source_cell.data_type != "f":
            return build_response(
                "error",
                None,
                exit_code=1,
                warnings=[f"Cell {args.cell} does not contain a formula"],
            )

        source_formula = source_cell.value

        # Copy down with reference adjustment
        for i in range(1, count + 1):
            target_row = source_row + i
            target_cell = ws.cell(row=target_row, column=source_col)

            # Adjust formula references (simple row adjustment)
            # This is a simplified implementation - real one would parse with openpyxl Tokenizer
            adjusted = _adjust_formula(source_formula, i)
            target_cell.value = adjusted
            formulas_copied += 1

    # Calculate filled range
    filled_range = f"{get_column_letter(source_col)}{source_row + 1}:{get_column_letter(source_col)}{source_row + count}"

    return build_response(
        "success",
        {
            "source": source,
            "target": args.target
            or f"{source}:{get_column_letter(source_col)}{source_row + count}",
            "filled_count": formulas_copied,
            "filled_range": filled_range,
        },
        impact={"cells_modified": formulas_copied, "formulas_added": formulas_copied},
        workbook_version=agent.version_hash,
    )


def _adjust_formula(formula: str, row_offset: int) -> str:
    """Simple formula adjustment - shifts relative row references."""
    import re

    # Pattern to match cell references (column + row)
    # This is simplified - real implementation would use openpyxl Tokenizer
    def shift_ref(match: re.Match) -> str:
        col = match.group(1)
        row = match.group(2)
        dollar = match.group(3) or ""

        # If absolute row ($), don't shift
        if dollar == "$":
            return match.group(0)

        new_row = int(row) + row_offset
        return f"{col}{dollar}{new_row}"

    # Match patterns like A1, $A1, A$1, $A$1
    pattern = r"([A-Z]+)(\$?)(\d+)"
    return re.sub(pattern, shift_ref, formula)


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()

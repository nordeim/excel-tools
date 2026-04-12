"""
Streaming I/O for large datasets in excel-agent-tools.

Provides chunked reading and row counting for worksheets exceeding
100k rows, keeping memory usage bounded by processing data in
configurable chunk sizes.

Uses openpyxl's iter_rows with values_only=True for efficient iteration.
The iter_rows generator doesn't load the entire sheet into memory.
"""

from __future__ import annotations

import datetime
from typing import Any, Generator

from openpyxl.worksheet.worksheet import Worksheet


def _serialize_cell_value(value: object) -> Any:  # noqa: ANN401
    """Convert a cell value to a JSON-serializable type.

    - datetime/date/time → ISO 8601 string
    - timedelta → total seconds as float
    - None → null
    - Everything else → passthrough (str, int, float, bool)
    """
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, datetime.time):
        return value.isoformat()
    if isinstance(value, datetime.timedelta):
        return value.total_seconds()
    return value


def read_range_chunked(
    sheet: Worksheet,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
    *,
    chunk_size: int = 10_000,
) -> Generator[list[list[Any]], None, None]:
    """Yield chunks of rows from a specified range.

    Each chunk is a list of rows, where each row is a list of cell values.
    Cell values are serialized to JSON-compatible types.

    Args:
        sheet: The worksheet to read from.
        min_row: First row (1-indexed).
        min_col: First column (1-indexed).
        max_row: Last row (1-indexed).
        max_col: Last column (1-indexed).
        chunk_size: Number of rows per chunk (default: 10,000).

    Yields:
        Lists of rows, each row being a list of serialized values.
    """
    current_chunk: list[list[Any]] = []

    for row_tuple in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        serialized = [_serialize_cell_value(v) for v in row_tuple]
        current_chunk.append(serialized)

        if len(current_chunk) >= chunk_size:
            yield current_chunk
            current_chunk = []

    if current_chunk:
        yield current_chunk


def read_range_full(
    sheet: Worksheet,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> list[list[Any]]:
    """Read an entire range into a single 2D list.

    Uses iter_rows with values_only for efficiency. Suitable for
    ranges that fit comfortably in memory (<100k rows).
    """
    result: list[list[Any]] = []
    for row_tuple in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        result.append([_serialize_cell_value(v) for v in row_tuple])
    return result


def count_used_rows(sheet: Worksheet) -> int:
    """Count rows that actually contain data.

    openpyxl's sheet.max_row can be inflated by formatting-only rows.
    This walks backwards from max_row to find the last row with data.
    """
    if sheet.max_row is None:
        return 0

    max_row = sheet.max_row
    max_col = sheet.max_column or 1

    for row_idx in range(max_row, 0, -1):
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                return row_idx
    return 0

"""
ExcelAgent: The central hub context manager for excel-agent-tools.

Integrates file locking, workbook loading, geometry hashing, and save
verification into a single, safe context manager.

Lifecycle:
__enter__:
1. Acquire FileLock (exclusive, with timeout)
2. Load workbook via openpyxl (preserving formulas and VBA)
3. Compute entry file hash (for concurrent modification detection)
4. Compute geometry hash (for version reporting to agent)

__exit__ (no exception, mode='rw'):
1. Re-read file hash from disk
2. If changed: raise ConcurrentModificationError (do NOT save)
3. Save workbook
4. Release lock

__exit__ (with exception):
1. Release lock WITHOUT saving (prevent partial/corrupt writes)
2. Re-raise the exception

Usage::

    with ExcelAgent(Path("workbook.xlsx"), mode="rw") as agent:
        ws = agent.workbook.active
        ws["A1"] = "Hello"
        # Saved automatically on __exit__ with hash verification
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from excel_agent.core.locking import FileLock
from excel_agent.core.version_hash import compute_file_hash, compute_workbook_hash
from excel_agent.utils.exceptions import (
    ConcurrentModificationError,
    ExcelFileNotFoundError,
    ValidationError,
)

if TYPE_CHECKING:
    from openpyxl import Workbook

logger = logging.getLogger(__name__)

_VALID_MODES = frozenset({"r", "rw"})
_VBA_EXTENSIONS = frozenset({".xlsm", ".xltm"})


class ExcelAgent:
    """Stateful context manager for safe, locked, hash-verified workbook manipulation.

    Args:
        path: Path to the Excel workbook (.xlsx, .xlsm, .xltx, .xltm).
        mode: "r" for read-only, "rw" for read-write (default).
        keep_vba: Preserve VBA projects in .xlsm files. Auto-detected from
            extension if not explicitly set.
        lock_timeout: Max seconds to wait for lock (default: 30).
        data_only: If True, read cached values instead of formulas.
            Default False — we ALWAYS preserve formulas.
    """

    def __init__(
        self,
        path: Path,
        *,
        mode: str = "rw",
        keep_vba: bool | None = None,
        lock_timeout: float = 30.0,
        data_only: bool = False,
    ) -> None:
        if mode not in _VALID_MODES:
            raise ValidationError(
                f"Invalid mode {mode!r}. Must be 'r' or 'rw'.",
                details={"mode": mode},
            )

        self._path = Path(path).resolve()
        self._mode = mode
        self._lock_timeout = lock_timeout
        self._data_only = data_only

        # Auto-detect VBA from extension if not explicitly set
        if keep_vba is None:
            self._keep_vba = self._path.suffix.lower() in _VBA_EXTENSIONS
        else:
            self._keep_vba = keep_vba

        # State (set during __enter__)
        self._lock: FileLock | None = None
        self._wb: Workbook | None = None
        self._entry_file_hash: str = ""
        self._geometry_hash: str = ""
        self._entered = False

    def __enter__(self) -> ExcelAgent:
        """Acquire lock, load workbook, compute hashes."""
        # Validate file exists
        if not self._path.exists():
            raise ExcelFileNotFoundError(
                f"Workbook not found: {self._path}",
                details={"path": str(self._path)},
            )
        if not self._path.is_file():
            raise ExcelFileNotFoundError(
                f"Path is not a file: {self._path}",
                details={"path": str(self._path)},
            )

        # Step 1: Acquire exclusive lock
        self._lock = FileLock(self._path, timeout=self._lock_timeout)
        self._lock.__enter__()

        try:
            # Step 2: Compute entry file hash (for concurrent modification detection)
            self._entry_file_hash = compute_file_hash(self._path)

            # Step 3: Load workbook
            self._wb = load_workbook(
                str(self._path),
                read_only=False,
                keep_vba=self._keep_vba,
                data_only=self._data_only,
                keep_links=True,
            )

            # Step 4: Compute geometry hash (for version reporting)
            self._geometry_hash = compute_workbook_hash(self._wb)

            self._entered = True
            logger.info(
                "ExcelAgent entered: %s (mode=%s, vba=%s, hash=%s)",
                self._path.name,
                self._mode,
                self._keep_vba,
                self._geometry_hash[:20] + "...",
            )
            return self

        except Exception:
            # If anything fails after lock acquisition, release the lock
            if self._lock is not None:
                self._lock.__exit__(None, None, None)
            self._lock = None
            raise

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: object,
    ) -> None:
        """Save (if rw + no exception), verify hash, release lock."""
        try:
            if exc_type is None and self._mode == "rw" and self._wb is not None:
                # Verify no concurrent modification before saving
                self.verify_no_concurrent_modification()
                # Save the workbook
                self._wb.save(str(self._path))
                logger.info("Workbook saved: %s", self._path.name)
            elif exc_type is not None:
                logger.warning(
                    "ExcelAgent exiting with exception — NOT saving: %s",
                    exc_type.__name__,
                )
        finally:
            # ALWAYS release the lock, even if save fails
            self._wb = None
            self._entered = False
            if self._lock is not None:
                self._lock.__exit__(None, None, None)
            self._lock = None

    @property
    def workbook(self) -> Workbook:
        """The openpyxl Workbook object.

        Raises:
            RuntimeError: If accessed outside the context manager.
        """
        if not self._entered or self._wb is None:
            raise RuntimeError(
                "ExcelAgent.workbook accessed outside context manager. "
                "Use 'with ExcelAgent(...) as agent:' to access the workbook."
            )
        return self._wb

    @property
    def path(self) -> Path:
        """The resolved workbook file path."""
        return self._path

    @property
    def version_hash(self) -> str:
        """The geometry hash computed at entry time.

        Returns:
            String in format "sha256:<hex_digest>"
        """
        return self._geometry_hash

    @property
    def file_hash(self) -> str:
        """The file byte hash computed at entry time."""
        return self._entry_file_hash

    def verify_no_concurrent_modification(self) -> None:
        """Re-read file bytes and compare hash to entry-time hash.

        Raises:
            ConcurrentModificationError: If the file was modified externally.
        """
        current_hash = compute_file_hash(self._path)
        if current_hash != self._entry_file_hash:
            raise ConcurrentModificationError(
                f"Workbook {self._path.name} was modified by another process "
                f"during this edit session. Refusing to save to prevent data loss.",
                details={
                    "path": str(self._path),
                    "entry_hash": self._entry_file_hash,
                    "current_hash": current_hash,
                },
            )

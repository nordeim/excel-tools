"""
Formula dependency graph engine for excel-agent-tools.

This is the most safety-critical component in the entire project. It powers
pre-flight impact reports that prevent AI agents from breaking formula chains.

Architecture:
1. Parse all formulas using openpyxl's Tokenizer
2. Extract cell references (Token.OPERAND + Token.RANGE subtype)
3. Build a directed graph: referenced_cell -> formula_cell
4. Provide BFS-based transitive closure for impact analysis
5. Detect circular references via Tarjan's SCC algorithm (iterative)

The Tokenizer identifies OPERAND tokens with RANGE subtype as cell/range
references. This includes direct references ($A$1, Sheet1!B2) AND named
ranges (DEFAULT_VAL). We use the Translator's CELL_REF_RE regex to
distinguish actual cell references from named ranges.

Performance target: 10-sheet, 1000-formula workbook analyzed in <5 seconds.
"""

from __future__ import annotations

import logging
import re
import time
from collections import deque
from collections.abc import Iterator
from dataclasses import dataclass, field
from typing import TYPE_CHECKING

from openpyxl.formula import Tokenizer
from openpyxl.formula.tokenizer import Token
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
)

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Regex patterns (borrowed from openpyxl.formula.translate.Translator)
# ---------------------------------------------------------------------------

# Matches a cell reference: A1, $A$1, AZ100, etc.
_CELL_REF_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]{0,6})$")

# Matches a range: A1:C10, $A$1:$C$10, etc.
_RANGE_REF_RE = re.compile(
    r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]{0,6})" r":\$?([A-Za-z]{1,3})\$?([1-9][0-9]{0,6})$"
)

# Matches an optional sheet prefix: Sheet1! or 'Sheet Name'!
_SHEET_PREFIX_RE = re.compile(r"^(?:'([^']+)'|([A-Za-z0-9_.\-]+))!(.+)$")


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class ImpactReport:
    """Pre-flight impact analysis for destructive operations.

    Returned by DependencyTracker.impact_report() to inform the AI agent
    about the consequences of a mutation before it happens.
    """

    status: str  # "safe" | "warning" | "critical"
    broken_references: int  # Number of formulas that would produce #REF!
    affected_sheets: list[str]  # Sheets containing affected formulas
    sample_errors: list[str]  # First 10 affected cells
    circular_refs_affected: bool  # Whether action affects a circular ref chain
    suggestion: str  # Prescriptive guidance for the agent
    details: dict[str, list[str]] = field(default_factory=dict)  # Per-sheet breakdown

    def to_dict(self) -> dict[str, object]:
        """Convert to JSON-serializable dict."""
        return {
            "status": self.status,
            "broken_references": self.broken_references,
            "affected_sheets": self.affected_sheets,
            "sample_errors": self.sample_errors,
            "circular_refs_affected": self.circular_refs_affected,
            "suggestion": self.suggestion,
            "details": self.details,
        }


# ---------------------------------------------------------------------------
# Reference extraction
# ---------------------------------------------------------------------------


def _normalize_cell_ref(ref_str: str, *, default_sheet: str) -> str:
    """Normalize a cell reference to 'SheetName!A1' canonical form.

    Strips $ anchors, resolves sheet prefix, uppercases column letters.

    Args:
        ref_str: Raw token value (e.g., "$A$1", "'Sheet 2'!B1", "A1:C10").
        default_sheet: Sheet name to use if no prefix is present.

    Returns:
        Normalized string like "Sheet1!A1" (for single cells).
        Returns None-equivalent empty string for unparseable references.
    """
    sheet = default_sheet
    ref = ref_str

    # Extract sheet prefix if present
    m = _SHEET_PREFIX_RE.match(ref)
    if m:
        sheet = m.group(1) or m.group(2)
        ref = m.group(3)

    return f"{sheet}!{ref.replace('$', '').upper()}"


def _expand_range_to_cells(normalized_ref: str) -> list[str]:
    """Expand a range reference into individual cell references.

    "Sheet1!A1:C3" -> ["Sheet1!A1", "Sheet1!A2", ..., "Sheet1!C3"]

    For single cells, returns a list with one element.
    For very large ranges (>10000 cells), returns just the corners
    to avoid memory explosion.
    """
    # Split sheet and range parts
    if "!" not in normalized_ref:
        return [normalized_ref]

    sheet, ref = normalized_ref.split("!", 1)

    # Try single cell first
    m = _CELL_REF_RE.match(ref)
    if m:
        return [normalized_ref]

    # Try range
    m = _RANGE_REF_RE.match(ref)
    if m:
        min_col = column_index_from_string(m.group(1))
        min_row = int(m.group(2))
        max_col = column_index_from_string(m.group(3))
        max_row = int(m.group(4))

        total_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        if total_cells > 10_000:
            # For very large ranges, track only the range itself as a unit
            # to avoid memory explosion. Impact analysis will still be correct
            # at the range level.
            logger.debug(
                "Range %s has %d cells — tracking as range unit", normalized_ref, total_cells
            )
            return [normalized_ref]

        cells: list[str] = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                col_letter = get_column_letter(col)
                cells.append(f"{sheet}!{col_letter}{row}")
        return cells

    # Unparseable - return as-is
    return [normalized_ref]


def _is_cell_reference(token_value: str) -> bool:
    """Check if a token value is an actual cell reference (not named range).

    Named ranges like "DEFAULT_VAL" also get RANGE subtype from Tokenizer,
    but they're not cell references. We use the regex to distinguish.
    """
    # Strip sheet prefix if present
    ref = token_value
    m = _SHEET_PREFIX_RE.match(ref)
    if m:
        ref = m.group(3)

    # Check if it matches cell or range pattern
    return bool(_CELL_REF_RE.match(ref) or _RANGE_REF_RE.match(ref))


# ---------------------------------------------------------------------------
# Tarjan's SCC Algorithm (iterative to avoid recursion limit)
# ---------------------------------------------------------------------------


def _tarjan_scc(graph: dict[str, set[str]]) -> list[list[str]]:
    """Find strongly connected components using Tarjan's algorithm (iterative).

    Returns cycles as lists of nodes. Each cycle is an SCC with >1 node
    or a self-loop (node pointing to itself).

    Time complexity: O(V + E)
    """
    index_counter = 0
    stack: list[str] = []
    lowlinks: dict[str, int] = {}
    index: dict[str, int] = {}
    on_stack: set[str] = set()
    sccs: list[list[str]] = []

    # Iterative DFS to avoid recursion limit
    for node in graph:
        if node in index:
            continue

        # Stack of (node, iterator_over_successors)
        dfs_stack: list[tuple[str, Iterator[str]]] = [(node, iter(graph.get(node, set())))]

        while dfs_stack:
            current, successors = dfs_stack[-1]

            if current not in index:
                # First time visiting this node
                index[current] = index_counter
                lowlinks[current] = index_counter
                index_counter += 1
                stack.append(current)
                on_stack.add(current)

            try:
                successor: str = next(successors)

                if successor not in index:
                    # Successor not yet visited - recurse
                    dfs_stack.append((successor, iter(graph.get(successor, set()))))
                elif successor in on_stack:
                    # Successor is on stack - update lowlink
                    lowlinks[current] = min(lowlinks[current], index[successor])

            except StopIteration:
                # Finished all successors of current
                dfs_stack.pop()

                # Propagate lowlink to parent
                if dfs_stack:
                    parent, _ = dfs_stack[-1]
                    lowlinks[parent] = min(lowlinks[parent], lowlinks[current])

                # Check if current is root of an SCC
                if lowlinks[current] == index[current]:
                    # Pop stack to form SCC
                    scc: list[str] = []
                    while True:
                        w = stack.pop()
                        on_stack.remove(w)
                        scc.append(w)
                        if w == current:
                            break

                    # Only include cycles (SCC with >1 node or self-loop)
                    if len(scc) > 1 or (len(scc) == 1 and current in graph.get(current, set())):
                        sccs.append(scc)

    return sccs


# ---------------------------------------------------------------------------
# Dependency Tracker
# ---------------------------------------------------------------------------


class DependencyTracker:
    """Builds and queries the workbook's formula dependency graph.

    This is the core safety component that prevents agents from breaking
    formula chains by providing pre-flight impact analysis.

    Usage:
        tracker = DependencyTracker(workbook)
        tracker.build_graph()

        # Check impact before deleting rows
        report = tracker.impact_report("Sheet1!A5:A10")
        if report.status != "safe":
            print(report.suggestion)
    """

    def __init__(self, workbook: Workbook):
        """Initialize with a workbook.

        Args:
            workbook: The openpyxl Workbook to analyze.
        """
        self._wb = workbook
        self._forward: dict[str, set[str]] = {}  # cell -> {cells that depend on it}
        self._reverse: dict[str, set[str]] = {}  # cell -> {cells it depends on}
        self._cycles: list[list[str]] = []
        self._formula_count = 0
        self._built = False
        self._build_time_ms = 0.0

    @property
    def is_built(self) -> bool:
        """Whether the graph has been built."""
        return self._built

    def build_graph(self, *, sheets: list[str] | None = None) -> None:
        """Parse all formulas and build the dependency graph.

        Uses openpyxl's Tokenizer to extract cell references from formulas.

        Args:
            sheets: Optional list of sheet names to analyze. If None, all sheets.
        """
        start_time = time.perf_counter()

        self._forward.clear()
        self._reverse.clear()
        self._formula_count = 0

        sheet_names = sheets or self._wb.sheetnames

        for sheet_name in sheet_names:
            if sheet_name not in self._wb.sheetnames:
                logger.warning("Sheet not found: %s", sheet_name)
                continue

            ws = self._wb[sheet_name]
            self._process_sheet(ws, sheet_name)

        # Detect circular references
        self._cycles = _tarjan_scc(self._reverse)

        self._built = True
        self._build_time_ms = (time.perf_counter() - start_time) * 1000

        logger.info(
            "Dependency graph built: %d formulas, %d edges, %d cycles, %.2fms",
            self._formula_count,
            sum(len(v) for v in self._forward.values()),
            len(self._cycles),
            self._build_time_ms,
        )

    def _process_sheet(self, ws: Worksheet, sheet_name: str) -> None:
        """Process all formulas in a sheet."""
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue  # Not a formula

                self._formula_count += 1
                formula_str = str(cell.value) if cell.value else ""

                # Parse formula to extract references
                self._process_formula(
                    formula_str,
                    cell.coordinate,
                    sheet_name,
                )

    def _process_formula(self, formula: str, cell_coord: str, sheet_name: str) -> None:
        """Extract references from a single formula."""
        try:
            tokenizer = Tokenizer(formula)
            cell_key = f"{sheet_name}!{cell_coord.replace('$', '').upper()}"

            for token in tokenizer.items:
                if token.type == Token.OPERAND and token.subtype == Token.RANGE:
                    ref_str = token.value

                    # Skip if not an actual cell reference (e.g., named range)
                    if not _is_cell_reference(ref_str):
                        continue

                    # Normalize and expand
                    normalized = _normalize_cell_ref(ref_str, default_sheet=sheet_name)
                    cells = _expand_range_to_cells(normalized)

                    for target_cell in cells:
                        # Add edge: target_cell -> cell_key
                        # (cell_key depends on target_cell)
                        if target_cell not in self._forward:
                            self._forward[target_cell] = set()
                        self._forward[target_cell].add(cell_key)

                        if cell_key not in self._reverse:
                            self._reverse[cell_key] = set()
                        self._reverse[cell_key].add(target_cell)

        except Exception as e:
            logger.warning("Failed to parse formula in %s!%s: %s", sheet_name, cell_coord, e)

    def find_dependents(self, target: str) -> set[str]:
        """Find all cells that would be affected if target is deleted/changed.

        Performs transitive closure (BFS) on the forward graph.
        A->B->C means deleting A affects B and C.

        Args:
            target: Cell reference like "Sheet1!A1" or "A1" (uses active sheet).

        Returns:
            Set of cell references that depend on the target.
        """
        if not self._built:
            raise RuntimeError("Graph not built. Call build_graph() first.")

        # Normalize target - preserve sheet name case, uppercase cell ref
        if "!" not in target:
            sheet = self._wb.sheetnames[0]
            ref = target.replace("$", "").upper()
            target = f"{sheet}!{ref}"
        else:
            parts = target.split("!", 1)
            sheet = parts[0].replace("$", "")
            ref = parts[1].replace("$", "").upper()
            target = f"{sheet}!{ref}"

        # BFS from target
        dependents: set[str] = set()
        queue: deque[str] = deque([target])
        visited: set[str] = {target}

        while queue:
            current = queue.popleft()

            for dependent in self._forward.get(current, set()):
                if dependent not in visited:
                    visited.add(dependent)
                    dependents.add(dependent)
                    queue.append(dependent)

        return dependents

    def find_precedents(self, cell: str) -> set[str]:
        """Find all cells that the given cell depends on (its inputs).

        Args:
            cell: Cell reference like "Sheet1!A1" or "A1".

        Returns:
            Set of cell references that the cell depends on.
        """
        if not self._built:
            raise RuntimeError("Graph not built. Call build_graph() first.")

        # Normalize cell - preserve sheet name case, uppercase cell ref
        if "!" not in cell:
            sheet = self._wb.sheetnames[0]
            ref = cell.replace("$", "").upper()
            cell = f"{sheet}!{ref}"
        else:
            parts = cell.split("!", 1)
            sheet = parts[0].replace("$", "")
            ref = parts[1].replace("$", "").upper()
            cell = f"{sheet}!{ref}"

        return self._reverse.get(cell, set()).copy()

    def impact_report(self, target_range: str, *, action: str = "delete") -> ImpactReport:
        """Generate a pre-flight impact report for a destructive operation.

        Args:
            target_range: Range that will be affected (e.g., "Sheet1!A5:A10").
            action: Type of operation ("delete", "insert", "modify").

        Returns:
            ImpactReport with status, broken references, and guidance.
        """
        if not self._built:
            raise RuntimeError("Graph not built. Call build_graph() first.")

        # Normalize and expand target range - preserve sheet name case
        if "!" not in target_range:
            sheet = self._wb.sheetnames[0]
            ref = target_range.replace("$", "").upper()
            normalized = f"{sheet}!{ref}"
        else:
            parts = target_range.split("!", 1)
            sheet = parts[0].replace("$", "")
            ref = parts[1].replace("$", "").upper()
            normalized = f"{sheet}!{ref}"

        target_cells = _expand_range_to_cells(normalized)

        # For very large ranges (sheet deletion), check each cell in the forward graph
        # that belongs to the target sheet. We detect this when _expand_range_to_cells
        # returns a single item that matches the normalized input exactly (meaning it
        # couldn't expand a range like A1:XFD1048576 due to size limit)
        if len(target_cells) == 1 and target_cells[0] == normalized:
            # Check if the normalized ref is a range pattern (contains ":")
            # and NOT a single cell
            if ":" in ref:
                # Range was too large to expand - check all cells in forward graph from this sheet
                target_cells = [
                    cell for cell in self._forward.keys() if cell.startswith(f"{sheet}!")
                ]
            # Otherwise it's a single cell, keep as-is
        all_dependents: set[str] = set()
        for cell in target_cells:
            all_dependents.update(self.find_dependents(cell))

        # Build per-sheet breakdown
        details: dict[str, list[str]] = {}
        for dependent in all_dependents:
            if "!" in dependent:
                sheet, coord = dependent.split("!", 1)
                if sheet not in details:
                    details[sheet] = []
                details[sheet].append(coord)

        affected_sheets = list(details.keys())
        broken_refs = len(all_dependents)

        # Check if any cycles are affected
        circular_affected = False
        for cycle in self._cycles:
            cycle_set = set(cycle)
            if any(cell in cycle_set for cell in target_cells):
                circular_affected = True
                break

        # Generate suggestion
        action_desc = {"delete": "deletion", "insert": "insertion", "modify": "modification"}.get(
            action, action
        )

        if broken_refs == 0:
            status = "safe"
            suggestion = "Operation is safe — no formulas will be affected."
        elif broken_refs <= 5:
            status = "warning"
            suggestion = (
                f"This {action_desc} will break {broken_refs} formula references. "
                "Consider using xls_update_references.py first."
            )
        else:
            status = "critical"
            suggestion = (
                f"CRITICAL: This {action_desc} will break {broken_refs} formula references "
                f"across {len(affected_sheets)} sheets. "
                f"Use xls_update_references.py --target='{target_range}' "
                "before proceeding."
            )

        # Append circular reference warning if applicable
        if circular_affected:
            suggestion += (
                " WARNING: This operation affects cells involved in circular reference chains. "
                "Review circular dependencies with xls-dependency-report before proceeding."
            )

        # Sample errors (first 10)
        sample_errors: list[str] = []
        for dependent in list(all_dependents)[:10]:
            sample_errors.append(f"{dependent} → #REF!")

        return ImpactReport(
            status=status,
            broken_references=broken_refs,
            affected_sheets=affected_sheets,
            sample_errors=sample_errors,
            circular_refs_affected=circular_affected,
            suggestion=suggestion,
            details=details,
        )

    def detect_circular_references(self) -> list[list[str]]:
        """Detect circular reference cycles in the workbook.

        Uses Tarjan's SCC algorithm (iterative implementation).

        Returns:
            List of cycles, where each cycle is a list of cell references.
        """
        if not self._built:
            raise RuntimeError("Graph not built. Call build_graph() first.")

        return [cycle.copy() for cycle in self._cycles]

    def get_adjacency_list(self) -> dict[str, list[str]]:
        """Export the dependency graph as a JSON-serializable adjacency list.

        Returns:
            Dict mapping cell references to lists of dependent cells.
        """
        if not self._built:
            raise RuntimeError("Graph not built. Call build_graph() first.")

        return {k: sorted(v) for k, v in self._forward.items()}

    def get_stats(self) -> dict[str, int]:
        """Get statistics about the dependency graph.

        Returns:
            Dict with total_cells, total_formulas, total_edges, circular_chains.
        """
        if not self._built:
            raise RuntimeError("Graph not built. Call build_graph() first.")

        return {
            "total_cells": len(self._forward),
            "total_formulas": self._formula_count,
            "total_edges": sum(len(v) for v in self._forward.values()),
            "circular_chains": len(self._cycles),
        }

"""
Geometry-aware workbook hashing for excel-agent-tools.

Produces SHA-256 hashes that capture workbook *structure and formulas*
but intentionally exclude cell values. This is because:
    - Values change on recalculation (they're volatile)
    - Formulas and structure are the "contract" (they're stable)
    - Concurrent modification detection needs to catch structural edits

Two hash types:
    - Geometry hash: sheet names + cell coordinates + formulas
    - File hash: raw bytes on disk (for concurrent modification detection)
"""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

_FILE_CHUNK_SIZE = 65536


def compute_workbook_hash(workbook: Workbook) -> str:
    """Compute a SHA-256 hash of workbook geometry.

    Includes:
        - Sheet names in order
        - Sheet visibility states
        - Cell coordinates that contain formulas
        - Formula strings

    Excludes:
        - Cell values (they change on recalculation)
        - Cell styles (they're presentational, not structural)
        - Workbook metadata (author, etc.)

    Returns:
        String in format "sha256:<hex_digest>"
    """
    h = hashlib.sha256()

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        h.update(f"SHEET:{sheet_name}".encode())
        h.update(f"STATE:{ws.sheet_state}".encode())
        _hash_sheet_geometry(ws, h)

    return f"sha256:{h.hexdigest()}"


def compute_sheet_hash(sheet: Worksheet) -> str:
    """Compute a SHA-256 hash of a single sheet's geometry.

    Returns:
        String in format "sha256:<hex_digest>"
    """
    h = hashlib.sha256()
    h.update(f"SHEET:{sheet.title}".encode())
    h.update(f"STATE:{sheet.sheet_state}".encode())
    _hash_sheet_geometry(sheet, h)
    return f"sha256:{h.hexdigest()}"


def compute_file_hash(path: Path) -> str:
    """Compute a SHA-256 hash of the raw file bytes on disk.

    This is used for concurrent modification detection: if the file
    changes on disk between our load and save, someone else modified it.

    Reads in 64KB chunks for memory efficiency on large files.

    Returns:
        String in format "sha256:<hex_digest>"
    """
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(_FILE_CHUNK_SIZE)
            if not chunk:
                break
            h.update(chunk)
    return f"sha256:{h.hexdigest()}"


def _hash_sheet_geometry(sheet: Worksheet, h: hashlib._Hash) -> None:
    """Hash the structural geometry of a single sheet.

    Iterates cells and includes coordinates + formulas for formula cells.
    Uses sorted iteration for deterministic ordering.
    """
    formula_cells: list[tuple[int, int, str]] = []

    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                formula_str = str(cell.value) if cell.value is not None else ""
                formula_cells.append((cell.row, cell.column, formula_str))

    formula_cells.sort()

    for row, col, formula in formula_cells:
        entry = f"F:{row}:{col}:{formula}"
        h.update(entry.encode("utf-8"))

    if sheet.min_row is not None:
        h.update(
            f"DIM:{sheet.min_row}:{sheet.min_column}:{sheet.max_row}:{sheet.max_column}".encode()
        )

"""
Unified range parsing for excel-agent-tools.

Converts any Excel reference format (A1, R1C1, Named Range, Table[Column])
into internal coordinates and back. This is the single source of truth for
all range resolution across the 53 tools.

Supported formats:
    - A1 notation: "A1", "A1:C10", "$A$1:$C$10", "Sheet1!A1:C10"
    - R1C1 notation: "R1C1", "R1C1:R10C3"
    - Named ranges: "SalesData" (requires workbook context)
    - Table references: "Table1[Sales]" (requires workbook context)
    - Full rows/columns: "A:A", "1:1"
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import TYPE_CHECKING

from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
)

from excel_agent.utils.exceptions import ValidationError

if TYPE_CHECKING:
    from openpyxl import Workbook


@dataclass(frozen=True)
class CellCoordinate:
    """A single cell coordinate (1-indexed)."""

    row: int
    col: int

    def __post_init__(self) -> None:
        if self.row < 1:
            raise ValidationError(f"Row must be >= 1, got {self.row}")
        if self.col < 1:
            raise ValidationError(f"Column must be >= 1, got {self.col}")


@dataclass(frozen=True)
class RangeCoordinate:
    """A rectangular cell range (1-indexed, inclusive).

    Attributes:
        sheet: Sheet name, or None for the active/default sheet.
        min_row: Top row (1-indexed).
        min_col: Left column (1-indexed).
        max_row: Bottom row, or None for single-cell / full-column.
        max_col: Right column, or None for single-cell / full-row.
    """

    sheet: str | None
    min_row: int
    min_col: int
    max_row: int | None = None
    max_col: int | None = None

    @property
    def is_single_cell(self) -> bool:
        return self.max_row is None and self.max_col is None


_CELL_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d{1,7})$")
_RANGE_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d{1,7}):\$?([A-Za-z]{1,3})\$?(\d{1,7})$")
_R1C1_CELL_RE = re.compile(r"^R(\d+)C(\d+)$", re.IGNORECASE)
_R1C1_RANGE_RE = re.compile(r"^R(\d+)C(\d+):R(\d+)C(\d+)$", re.IGNORECASE)
_FULL_COL_RE = re.compile(r"^\$?([A-Za-z]{1,3}):\$?([A-Za-z]{1,3})$")
_FULL_ROW_RE = re.compile(r"^(\d+):(\d+)$")
_SHEET_PREFIX_RE = re.compile(r"^(?:'([^']+)'|([A-Za-z0-9_.\-]+))!(.+)$")
_TABLE_REF_RE = re.compile(r"^([A-Za-z_][A-Za-z0-9_.]*)(?:\[([^\]]*)\])?$")


def col_letter_to_number(letter: str) -> int:
    """Convert column letter(s) to 1-indexed number."""
    return column_index_from_string(letter.upper())


def col_number_to_letter(number: int) -> str:
    """Convert 1-indexed column number to letter(s)."""
    return get_column_letter(number)


class RangeSerializer:
    """Parses any supported Excel reference format to RangeCoordinate.

    Args:
        workbook: Optional workbook context for resolving named ranges
                  and table references. If None, only A1/R1C1 formats
                  are supported.
    """

    def __init__(self, workbook: Workbook | None = None) -> None:
        self._wb = workbook

    def parse(self, range_str: str, *, default_sheet: str | None = None) -> RangeCoordinate:
        """Parse any supported format to RangeCoordinate.

        Args:
            range_str: The reference string (A1, R1C1, named range, table ref).
            default_sheet: Sheet name to use if not specified in range_str.

        Returns:
            Parsed RangeCoordinate.

        Raises:
            ValidationError: If the input cannot be parsed.
        """
        original = range_str.strip()
        if not original:
            raise ValidationError("Empty range string")

        sheet: str | None = default_sheet
        ref = original
        m = _SHEET_PREFIX_RE.match(ref)
        if m:
            sheet = m.group(1) or m.group(2)
            ref = m.group(3)

        # A1 range
        m = _RANGE_RE.match(ref)
        if m:
            return RangeCoordinate(
                sheet=sheet,
                min_row=int(m.group(2)),
                min_col=col_letter_to_number(m.group(1)),
                max_row=int(m.group(4)),
                max_col=col_letter_to_number(m.group(3)),
            )

        # A1 single cell
        m = _CELL_RE.match(ref)
        if m:
            return RangeCoordinate(
                sheet=sheet,
                min_row=int(m.group(2)),
                min_col=col_letter_to_number(m.group(1)),
            )

        # R1C1 range
        m = _R1C1_RANGE_RE.match(ref)
        if m:
            return RangeCoordinate(
                sheet=sheet,
                min_row=int(m.group(1)),
                min_col=int(m.group(2)),
                max_row=int(m.group(3)),
                max_col=int(m.group(4)),
            )

        # R1C1 single cell
        m = _R1C1_CELL_RE.match(ref)
        if m:
            return RangeCoordinate(
                sheet=sheet,
                min_row=int(m.group(1)),
                min_col=int(m.group(2)),
            )

        # Full column
        m = _FULL_COL_RE.match(ref)
        if m:
            return RangeCoordinate(
                sheet=sheet,
                min_row=1,
                min_col=col_letter_to_number(m.group(1)),
                max_row=None,
                max_col=col_letter_to_number(m.group(2)),
            )

        # Full row
        m = _FULL_ROW_RE.match(ref)
        if m:
            return RangeCoordinate(
                sheet=sheet,
                min_row=int(m.group(1)),
                min_col=1,
                max_row=int(m.group(2)),
                max_col=None,
            )

        # Named range
        if self._wb is not None:
            coord = self._try_named_range(ref, sheet)
            if coord is not None:
                return coord

        # Table reference
        if self._wb is not None:
            m = _TABLE_REF_RE.match(ref)
            if m:
                coord = self._try_table_ref(m.group(1), m.group(2), sheet)
                if coord is not None:
                    return coord

        raise ValidationError(
            f"Cannot parse range: {original!r}. "
            "Expected A1, R1C1, named range, or table reference.",
            details={"input": original},
        )

    def _try_named_range(self, name: str, fallback_sheet: str | None) -> RangeCoordinate | None:
        """Attempt to resolve a named range from the workbook."""
        if self._wb is None:
            return None

        for defn_name, defn in self._wb.defined_names.items():
            if defn_name.lower() == name.lower():
                destinations = list(defn.destinations)
                if not destinations:
                    return None
                dest_sheet, dest_range = destinations[0]
                return self.parse(
                    dest_range,
                    default_sheet=dest_sheet or fallback_sheet,
                )
        return None

    def _try_table_ref(
        self,
        table_name: str,
        column_spec: str | None,
        fallback_sheet: str | None,
    ) -> RangeCoordinate | None:
        """Attempt to resolve a table reference from the workbook."""
        if self._wb is None:
            return None

        for ws in self._wb.worksheets:
            for table in ws.tables.values():
                if table.name.lower() == table_name.lower():
                    sheet_name = ws.title
                    ref = table.ref
                    if column_spec is None:
                        return self.parse(ref, default_sheet=sheet_name)
                    for idx, col in enumerate(table.tableColumns):
                        if col.name.lower() == column_spec.lower():
                            parsed = self.parse(ref, default_sheet=sheet_name)
                            target_col = parsed.min_col + idx
                            return RangeCoordinate(
                                sheet=sheet_name,
                                min_row=parsed.min_row,
                                min_col=target_col,
                                max_row=parsed.max_row,
                                max_col=target_col,
                            )
                    return None
        return None

    @staticmethod
    def to_a1(coord: RangeCoordinate) -> str:
        """Convert RangeCoordinate to A1 notation string."""
        prefix = ""
        if coord.sheet is not None:
            if re.search(r"[^A-Za-z0-9_.]", coord.sheet):
                prefix = f"'{coord.sheet}'!"
            else:
                prefix = f"{coord.sheet}!"

        start = f"{col_number_to_letter(coord.min_col)}{coord.min_row}"

        if coord.max_row is None and coord.max_col is None:
            return f"{prefix}{start}"

        if coord.max_row is not None and coord.max_col is not None:
            end = f"{col_number_to_letter(coord.max_col)}{coord.max_row}"
            return f"{prefix}{start}:{end}"

        if coord.max_row is not None and coord.max_col is None:
            return f"{prefix}{coord.min_row}:{coord.max_row}"
        return (
            f"{prefix}{col_number_to_letter(coord.min_col)}:{col_number_to_letter(coord.max_col)}"
        )

    @staticmethod
    def to_r1c1(coord: RangeCoordinate) -> str:
        """Convert RangeCoordinate to R1C1 notation string."""
        start = f"R{coord.min_row}C{coord.min_col}"
        if coord.max_row is None and coord.max_col is None:
            return start
        if coord.max_row is not None and coord.max_col is not None:
            return f"{start}:R{coord.max_row}C{coord.max_col}"
        return start

"""
Cell style serialization for excel-agent-tools.

Converts openpyxl style objects (Font, PatternFill, Border, Alignment)
into JSON-serializable dicts for the xls_get_cell_style tool.

openpyxl Color objects can be indexed, themed, or aRGB. We normalize
everything to hex strings where possible. Per openpyxl docs:
"It is advisable to use aRGB colours."
"""

from __future__ import annotations

from typing import Any

from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color


def _serialize_color(color: Color | None) -> str | None:
    """Convert an openpyxl Color to a hex string or None."""
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        rgb = str(color.rgb)
        # Strip alpha prefix if present (AARRGGBB → RRGGBB)
        if len(rgb) == 8:
            return rgb[2:]
        return rgb
    if color.type == "theme":
        return f"theme:{color.theme}"
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    return None


def _serialize_side(side: Side | None) -> dict[str, Any] | None:
    """Serialize a border Side."""
    if side is None:
        return None
    if side.border_style is None:
        return None
    result: dict[str, Any] = {"style": side.border_style}
    if side.color:
        result["color"] = _serialize_color(side.color)
    return result


def serialize_font(font: Font) -> dict[str, Any]:
    """Serialize an openpyxl Font to a dict."""
    result: dict[str, Any] = {}
    if font.name:
        result["name"] = font.name
    if font.size is not None:
        result["size"] = font.size
    if font.bold is not None:
        result["bold"] = font.bold
    if font.italic is not None:
        result["italic"] = font.italic
    if font.underline and font.underline != "none":
        result["underline"] = font.underline
    if font.strike is not None:
        result["strikethrough"] = font.strike
    if font.color:
        result["color"] = _serialize_color(font.color)
    if font.vertAlign:
        result["vertAlign"] = font.vertAlign
    return result


def serialize_fill(fill: PatternFill) -> dict[str, Any]:
    """Serialize an openpyxl PatternFill to a dict."""
    result: dict[str, Any] = {}
    if fill.fill_type:
        result["patternType"] = fill.fill_type
    if fill.fgColor:
        result["fgColor"] = _serialize_color(fill.fgColor)
    if fill.bgColor:
        result["bgColor"] = _serialize_color(fill.bgColor)
    return result


def serialize_border(border: Border) -> dict[str, Any]:
    """Serialize an openpyxl Border to a dict."""
    result: dict[str, Any] = {}
    for side_name in ("top", "bottom", "left", "right"):
        side = getattr(border, side_name, None)
        serialized = _serialize_side(side)
        if serialized:
            result[side_name] = serialized
    return result


def serialize_alignment(alignment: Alignment) -> dict[str, Any]:
    """Serialize an openpyxl Alignment to a dict."""
    result: dict[str, Any] = {}
    if alignment.horizontal:
        result["horizontal"] = alignment.horizontal
    if alignment.vertical:
        result["vertical"] = alignment.vertical
    if alignment.text_rotation:
        result["textRotation"] = alignment.text_rotation
    if alignment.wrap_text is not None:
        result["wrapText"] = alignment.wrap_text
    if alignment.shrink_to_fit is not None:
        result["shrinkToFit"] = alignment.shrink_to_fit
    return result


def serialize_cell_style(cell: Cell) -> dict[str, Any]:
    """Serialize all style properties of a cell to a JSON dict."""
    return {
        "font": serialize_font(cell.font),
        "fill": serialize_fill(cell.fill),
        "border": serialize_border(cell.border),
        "alignment": serialize_alignment(cell.alignment),
        "number_format": cell.number_format,
    }

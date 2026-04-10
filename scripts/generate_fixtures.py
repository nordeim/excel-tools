#!/usr/bin/env python3
"""
Generate realistic office fixtures for E2E testing.

Creates:
1. OfficeOps_Expenses_KPI.xlsx - Main "real life" workbook
2. EdgeCases_Formulas_and_Links.xlsx - Edge case formulas
3. Macro binaries (safe.bin, risky.bin) - Macro artifacts

Usage:
    python scripts/generate_fixtures.py
"""

import os
import random
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Set seed for deterministic generation
random.seed(42)

# Ensure fixtures directory exists
FIXTURES_DIR = Path(__file__).parent.parent / "tests" / "fixtures"
FIXTURES_DIR.mkdir(parents=True, exist_ok=True)


def generate_officeops_expenses_kpi():
    """Generate Fixture 1: OfficeOps_Expenses_KPI.xlsx"""
    wb = Workbook()

    # Sheet 1: Lists (reference data)
    ws_lists = wb.active
    ws_lists.title = "Lists"

    # Categories in A1:A10
    categories = [
        "Travel",
        "Software",
        "Meals",
        "Shipping",
        "Contractor",
        "OfficeSupplies",
        "Marketing",
        "Cloud",
        "Training",
        "Other",
    ]
    for i, cat in enumerate(categories, 1):
        ws_lists.cell(row=i, column=1, value=cat)

    # Departments in B1:B6
    departments = ["Sales", "Finance", "Ops", "Eng", "HR", "Exec"]
    for i, dept in enumerate(departments, 1):
        ws_lists.cell(row=i, column=2, value=dept)

    # TaxRate in D2
    ws_lists.cell(row=2, column=4, value=0.0825)
    ws_lists.cell(row=2, column=3, value="TaxRate")

    # ReportMonth in D3
    ws_lists.cell(row=3, column=4, value=datetime(2026, 3, 1))
    ws_lists.cell(row=3, column=3, value="ReportMonth")

    # Define named ranges
    wb.create_named_range("Categories", ws_lists, "$A$1:$A$10")
    wb.create_named_range("Departments", ws_lists, "$B$1:$B$6")
    wb.create_named_range("TaxRate", ws_lists, "$D$2")
    wb.create_named_range("ReportMonth", ws_lists, "$D$3")

    # Sheet 2: Raw_Expenses
    ws_expenses = wb.create_sheet("Raw_Expenses")

    # Headers
    headers = [
        "Date",
        "Dept",
        "Vendor",
        "Category",
        "Amount",
        "Currency",
        "FX",
        "AmountUSD",
        "Notes",
        "ReceiptURL",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws_expenses.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Generate 200 rows of expense data
    vendors = ["VendorA", "VendorB", "VendorC", "Amazon", "Microsoft", "Oracle", "SalesForce"]
    currencies = ["USD", "EUR", "GBP", "JPY"]
    notes_samples = [
        "Regular purchase",
        'Special order with "quotes" in description',
        "Multi-line\nnote with\nnewlines",
        "Comma, separated, values",
        "Normal note",
    ]

    for row in range(2, 202):
        # Date (spread over March 2026)
        date_offset = random.randint(0, 30)
        ws_expenses.cell(
            row=row, column=1, value=datetime(2026, 3, 1) + timedelta(days=date_offset)
        )

        # Dept (random from departments)
        ws_expenses.cell(row=row, column=2, value=random.choice(departments))

        # Vendor
        ws_expenses.cell(row=row, column=3, value=random.choice(vendors))

        # Category
        ws_expenses.cell(row=row, column=4, value=random.choice(categories))

        # Amount (100-5000)
        ws_expenses.cell(row=row, column=5, value=round(random.uniform(100, 5000), 2))

        # Currency
        ws_expenses.cell(row=row, column=6, value=random.choice(currencies))

        # FX formula (will be =IF(F2="USD",1,XLOOKUP(...)))
        # For now, leave blank - will be set by xls-set-formula

        # AmountUSD formula (will be =E2*G2)
        # For now, leave blank - will be set by xls-set-formula

        # Notes
        ws_expenses.cell(row=row, column=9, value=random.choice(notes_samples))

        # ReceiptURL
        ws_expenses.cell(row=row, column=10, value=f"https://receipts.example.com/{row:06d}")

    # Add one deliberate error row at the end
    ws_expenses.cell(
        row=202, column=1, value='=VLOOKUP("missing",Lists!A:B,2,FALSE)'
    )  # Will cause #N/A

    # Add data validation for Dept column (B)
    dv_dept = DataValidation(type="list", formula1="=Departments", allow_blank=False)
    dv_dept.error = "Please select a valid department"
    dv_dept.prompt = "Select Department"
    ws_expenses.add_data_validation(dv_dept)
    dv_dept.add(f"B2:B201")

    # Add data validation for Category column (D)
    dv_cat = DataValidation(type="list", formula1="=Categories", allow_blank=False)
    dv_cat.error = "Please select a valid category"
    dv_cat.prompt = "Select Category"
    ws_expenses.add_data_validation(dv_cat)
    dv_cat.add(f"D2:D201")

    # Sheet 3: FXRates
    ws_fx = wb.create_sheet("FXRates")
    ws_fx.cell(row=1, column=1, value="Currency")
    ws_fx.cell(row=1, column=2, value="Rate")

    fx_rates = [
        ("USD", 1.0),
        ("EUR", 1.10),
        ("GBP", 1.30),
        ("JPY", 0.009),
    ]
    for i, (curr, rate) in enumerate(fx_rates, 2):
        ws_fx.cell(row=i, column=1, value=curr)
        ws_fx.cell(row=i, column=2, value=rate)

    # Sheet 4: Summary
    ws_summary = wb.create_sheet("Summary")

    # KPI cells
    ws_summary.cell(row=1, column=1, value="Total Spend:")
    ws_summary.cell(row=1, column=2, value="=SUM(Raw_Expenses[AmountUSD])")  # Structured reference

    ws_summary.cell(row=2, column=1, value="Spend by Department")

    # Department spend matrix
    ws_summary.cell(row=3, column=1, value="Department")
    ws_summary.cell(row=3, column=2, value="Total")

    for i, dept in enumerate(departments, 4):
        ws_summary.cell(row=i, column=1, value=dept)
        ws_summary.cell(
            row=i, column=2, value=f'=SUMIF(Raw_Expenses[Dept],"{dept}",Raw_Expenses[AmountUSD])'
        )

    # Month filter
    ws_summary.cell(row=10, column=1, value="Report Month:")
    ws_summary.cell(row=10, column=2, value="=EOMONTH(ReportMonth,0)")

    # Sheet 5: Dashboard
    ws_dash = wb.create_sheet("Dashboard")

    # Merged title cell
    ws_dash.merge_cells("A1:D1")
    title_cell = ws_dash.cell(row=1, column=1, value="Expense Dashboard - March 2026")
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Reserve space for charts
    ws_dash.cell(row=3, column=7, value="Chart Area")

    # Freeze panes
    ws_dash.freeze_panes = "A3"

    # Save
    output_path = FIXTURES_DIR / "OfficeOps_Expenses_KPI.xlsx"
    wb.save(output_path)
    print(f"✅ Generated: {output_path}")

    return output_path


def generate_edge_cases_formulas():
    """Generate Fixture 2: EdgeCases_Formulas_and_Links.xlsx"""
    wb = Workbook()

    # Sheet 1: Circular references
    ws_circ = wb.active
    ws_circ.title = "Circular"
    ws_circ.cell(row=1, column=1, value="=B1+1")
    ws_circ.cell(row=1, column=2, value="=A1+1")

    # Sheet 2: Dynamic Arrays (these will likely fail Tier 1)
    ws_dyn = wb.create_sheet("DynamicArrays")
    ws_dyn.cell(row=1, column=1, value="=UNIQUE(Range1)")
    ws_dyn.cell(row=2, column=1, value="=FILTER(Range1, Condition)")
    ws_dyn.cell(row=3, column=1, value="=LET(x, 5, y, 10, x + y)")
    ws_dyn.cell(row=4, column=1, value='=TEXTSPLIT("hello world", " ")')

    # Sheet 3: External Links
    ws_ext = wb.create_sheet("ExternalLinks")
    ws_ext.cell(row=1, column=1, value="External Reference")
    ws_ext.cell(row=2, column=1, value="='[OtherBook.xlsx]Sheet1'!A1")

    output_path = FIXTURES_DIR / "EdgeCases_Formulas_and_Links.xlsx"
    wb.save(output_path)
    print(f"✅ Generated: {output_path}")

    return output_path


def generate_macro_binaries():
    """Generate macro binaries (safe.bin and risky.bin)"""
    # Create directories
    macros_dir = FIXTURES_DIR / "macros"
    macros_dir.mkdir(exist_ok=True)

    # Generate safe VBA project (benign formatting code)
    safe_bin = macros_dir / "vbaProject_safe.bin"
    # This is a minimal VBA project header + benign code
    safe_content = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # Compound file header
    safe_content += b"Module=SafeModule\x00"
    safe_content += b'Attribute VB_Name = "SafeModule"\n'
    safe_content += b"Sub FormatNumbers()\n"
    safe_content += b"    Dim cell As Range\n"
    safe_content += b"    For Each cell In Selection\n"
    safe_content += b'        cell.NumberFormat = "#,##0.00"\n'
    safe_content += b"    Next cell\n"
    safe_content += b"End Sub\n"
    safe_content += b"\x00" * 100  # Padding

    with open(safe_bin, "wb") as f:
        f.write(safe_content)
    print(f"✅ Generated: {safe_bin}")

    # Generate risky VBA project (contains suspicious patterns)
    risky_bin = macros_dir / "vbaProject_risky.bin"
    risky_content = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # Compound file header
    risky_content += b"Module=RiskyModule\x00"
    risky_content += b'Attribute VB_Name = "RiskyModule"\n'
    risky_content += b"Sub AutoOpen()\n"  # Auto-execute
    risky_content += b"    Dim url As String\n"
    risky_content += b"    url = Chr(104) & Chr(116) & Chr(116) & Chr(112)\n"  # Obfuscated
    risky_content += (
        b'    Shell "powershell -Command ""Invoke-WebRequest -Uri "" & url""", vbHide\n'  # Shell
    )
    risky_content += b"End Sub\n"
    risky_content += b"\x00" * 100  # Padding

    with open(risky_bin, "wb") as f:
        f.write(risky_content)
    print(f"✅ Generated: {risky_bin}")

    # Generate MacroTarget.xlsx (normal xlsx to inject into)
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(row=1, column=1, value="Target Data")
    ws.cell(row=2, column=1, value=100)
    ws.cell(row=3, column=1, value=200)

    target_path = FIXTURES_DIR / "MacroTarget.xlsx"
    wb.save(target_path)
    print(f"✅ Generated: {target_path}")

    return safe_bin, risky_bin, target_path


def main():
    """Generate all fixtures"""
    print("=" * 70)
    print("Generating Realistic Office Fixtures")
    print("=" * 70)
    print()

    # Generate fixtures
    office_fixture = generate_officeops_expenses_kpi()
    edge_fixture = generate_edge_cases_formulas()
    safe_bin, risky_bin, target = generate_macro_binaries()

    print()
    print("=" * 70)
    print("Fixture Generation Complete")
    print("=" * 70)
    print()
    print("Generated files:")
    print(f"  1. {office_fixture.name}")
    print(f"  2. {edge_fixture.name}")
    print(f"  3. {safe_bin.name}")
    print(f"  4. {risky_bin.name}")
    print(f"  5. {target.name}")
    print()
    print(f"All fixtures saved to: {FIXTURES_DIR}")


if __name__ == "__main__":
    main()

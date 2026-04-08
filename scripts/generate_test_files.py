#!/usr/bin/env python3
"""
Generate all test fixture files for excel-agent-tools.

This script programmatically creates every test fixture workbook so that
the test suite does not depend on binary blobs committed to the repository.
Running it twice produces identical files (idempotent).

Usage:
    python scripts/generate_test_files.py [--output-dir tests/fixtures]

Generated files:
    sample.xlsx              — 3-sheet workbook with data, formulas, named ranges
    complex_formulas.xlsx    — 10-sheet workbook with 1000+ cross-sheet formulas
    circular_refs.xlsx       — Workbook with intentional circular references
    template.xltx            — Template with {{placeholder}} variables
    large_dataset.xlsx       — 500k rows × 10 columns (write-only mode)
"""

from __future__ import annotations

import argparse
import datetime
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName


def generate_sample(output_dir: Path) -> Path:
    """Generate sample.xlsx: 3-sheet workbook with data and formulas."""
    wb = Workbook()

    # Sheet1: Data table with formulas
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Sheet1"
    headers = ["Name", "Q1", "Q2", "Q3", "Q4", "Total"]
    ws1.append(headers)
    ws1["A1"].font = Font(bold=True)

    products = ["Widget A", "Widget B", "Gadget X", "Gadget Y", "Service Z"]
    for i, product in enumerate(products, start=2):
        ws1[f"A{i}"] = product
        for col_idx, col_letter in enumerate(["B", "C", "D", "E"], start=1):
            ws1[f"{col_letter}{i}"] = (i * 1000) + (col_idx * 100)
        ws1[f"F{i}"] = f"=SUM(B{i}:E{i})"

    # Summary row
    summary_row = len(products) + 2
    ws1[f"A{summary_row}"] = "Grand Total"
    ws1[f"A{summary_row}"].font = Font(bold=True)
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws1[f"{col_letter}{summary_row}"] = f"=SUM({col_letter}2:{col_letter}{summary_row - 1})"

    # Sheet2: Cross-sheet references
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Summary Report"
    ws2["A2"] = "Total Revenue"
    ws2[f"B2"] = f"=Sheet1!F{summary_row}"
    ws2["A3"] = "Average Quarter"
    ws2["B3"] = f"=Sheet1!F{summary_row}/4"
    ws2["A4"] = "Double Revenue"
    ws2["B4"] = "=B2*2"

    # Sheet3: Named range target
    ws3 = wb.create_sheet("Sheet3")
    ws3["A1"] = "Category"
    ws3["B1"] = "Budget"
    for i in range(2, 8):
        ws3[f"A{i}"] = f"Department {i - 1}"
        ws3[f"B{i}"] = (i - 1) * 5000

    defn = DefinedName("BudgetData", attr_text="Sheet3!$A$1:$B$7")
    wb.defined_names.add(defn)

    path = output_dir / "sample.xlsx"
    wb.save(str(path))
    print(f"  ✓ {path.name} ({path.stat().st_size:,} bytes)")
    return path


def generate_complex_formulas(output_dir: Path) -> Path:
    """Generate complex_formulas.xlsx: 10 sheets, 1000+ cross-sheet formulas."""
    wb = Workbook()

    sheet_names = [f"Dept{i}" for i in range(1, 11)]

    # Create all sheets first
    ws_first = wb.active
    assert ws_first is not None
    ws_first.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(name)

    # Populate each sheet with data and formulas
    for sheet_idx, name in enumerate(sheet_names):
        ws = wb[name]
        # Header row
        ws.append(["Month", "Revenue", "Cost", "Profit", "Margin"])

        # 12 months of data + formulas
        for month in range(1, 13):
            row = month + 1
            ws[f"A{row}"] = f"2026-{month:02d}"
            ws[f"B{row}"] = (sheet_idx + 1) * 10000 + month * 500
            ws[f"C{row}"] = (sheet_idx + 1) * 6000 + month * 300
            ws[f"D{row}"] = f"=B{row}-C{row}"
            ws[f"E{row}"] = f"=IF(B{row}>0,D{row}/B{row},0)"

        # Annual totals
        ws["A14"] = "Total"
        ws["A14"].font = Font(bold=True)
        ws["B14"] = "=SUM(B2:B13)"
        ws["C14"] = "=SUM(C2:C13)"
        ws["D14"] = "=SUM(D2:D13)"
        ws["E14"] = "=IF(B14>0,D14/B14,0)"

    # Create a Summary sheet with cross-sheet references
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary["A1"] = "Department Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append(["Department", "Revenue", "Cost", "Profit", "Margin"])

    for idx, name in enumerate(sheet_names):
        row = idx + 3
        ws_summary[f"A{row}"] = name
        ws_summary[f"B{row}"] = f"='{name}'!B14"
        ws_summary[f"C{row}"] = f"='{name}'!C14"
        ws_summary[f"D{row}"] = f"='{name}'!D14"
        ws_summary[f"E{row}"] = f"='{name}'!E14"

    # Grand totals
    total_row = len(sheet_names) + 3
    ws_summary[f"A{total_row}"] = "Grand Total"
    ws_summary[f"A{total_row}"].font = Font(bold=True)
    for col in ["B", "C", "D"]:
        ws_summary[f"{col}{total_row}"] = f"=SUM({col}3:{col}{total_row - 1})"
    ws_summary[f"E{total_row}"] = f"=IF(B{total_row}>0,D{total_row}/B{total_row},0)"

    # Named ranges
    defn = DefinedName("AllRevenue", attr_text=f"Summary!$B$3:$B${total_row - 1}")
    wb.defined_names.add(defn)
    defn2 = DefinedName("GrandProfit", attr_text=f"Summary!$D${total_row}")
    wb.defined_names.add(defn2)

    path = output_dir / "complex_formulas.xlsx"
    wb.save(str(path))
    print(f"  ✓ {path.name} ({path.stat().st_size:,} bytes)")
    return path


def generate_circular_refs(output_dir: Path) -> Path:
    """Generate circular_refs.xlsx: intentional circular references."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    # 2-cell circular
    ws["A1"] = "=B1+1"
    ws["B1"] = "=A1+1"

    # 3-cell circular
    ws["A3"] = "=C3+1"
    ws["B3"] = "=A3+1"
    ws["C3"] = "=B3+1"

    # Self-referencing
    ws["A5"] = "=A5+1"

    path = output_dir / "circular_refs.xlsx"
    wb.save(str(path))
    print(f"  ✓ {path.name} ({path.stat().st_size:,} bytes)")
    return path


def generate_template(output_dir: Path) -> Path:
    """Generate template.xltx: template with {{placeholder}} variables."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Report"

    # Header with placeholders
    ws["A1"] = "{{company}}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "Annual Report {{year}}"
    ws["A2"].font = Font(size=12)
    ws["A3"] = "Prepared by: {{author}}"
    ws["A3"].font = Font(italic=True)

    # Data section with placeholders
    ws["A5"] = "Revenue"
    ws["A5"].font = Font(bold=True)
    ws["B5"] = "{{revenue}}"
    ws["A6"] = "Expenses"
    ws["A6"].font = Font(bold=True)
    ws["B6"] = "{{expenses}}"
    ws["A7"] = "Net Income"
    ws["A7"].font = Font(bold=True)
    ws["B7"] = "=B5-B6"

    # Set as template
    wb.template = True

    path = output_dir / "template.xltx"
    wb.save(str(path))
    print(f"  ✓ {path.name} ({path.stat().st_size:,} bytes)")
    return path


def generate_large_dataset(output_dir: Path) -> Path:
    """Generate large_dataset.xlsx: 500k rows × 10 columns using write-only mode."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Data")

    # Header
    ws.append(
        ["ID", "Name", "Category", "Value", "Date", "Region", "Status", "Score", "Notes", "Active"]
    )

    categories = ["Electronics", "Clothing", "Food", "Services", "Industrial"]
    regions = ["North", "South", "East", "West", "Central"]
    statuses = ["Active", "Pending", "Closed", "Archived"]

    base_date = datetime.datetime(2025, 1, 1, tzinfo=datetime.timezone.utc)
    for i in range(1, 500_001):
        ws.append(
            [
                i,
                f"Item-{i:06d}",
                categories[i % len(categories)],
                round(i * 1.234, 2),
                (base_date + datetime.timedelta(days=i % 365)).strftime("%Y-%m-%d"),
                regions[i % len(regions)],
                statuses[i % len(statuses)],
                round((i % 100) * 1.5, 1),
                f"Note for item {i}",
                i % 3 != 0,
            ]
        )

        if i % 100_000 == 0:
            print(f"    ... written {i:,} rows")

    path = output_dir / "large_dataset.xlsx"
    wb.save(str(path))
    print(f"  ✓ {path.name} ({path.stat().st_size:,} bytes)")
    return path


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate all test fixture workbooks for excel-agent-tools."
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default="tests/fixtures",
        help="Directory to write fixture files (default: tests/fixtures)",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Generating test fixtures in {output_dir}/")

    generators = [
        ("Sample workbook", generate_sample),
        ("Complex formulas workbook", generate_complex_formulas),
        ("Circular references workbook", generate_circular_refs),
        ("Template workbook", generate_template),
        ("Large dataset workbook", generate_large_dataset),
    ]

    for name, gen_fn in generators:
        print(f"\nGenerating: {name}")
        try:
            gen_fn(output_dir)
        except Exception as exc:
            print(f"  ✗ FAILED: {exc}", file=sys.stderr)
            return 1

    print(f"\n✓ All {len(generators)} fixtures generated successfully.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

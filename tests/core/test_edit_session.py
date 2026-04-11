"""Unit tests for EditSession abstraction.

Tests the unified edit target semantics that eliminate double-save issues.
"""

from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_agent.core.edit_session import EditSession, prepare_edit_target
from excel_agent.utils.exceptions import ValidationError


class TestEditSessionPrepare:
    """Tests for EditSession.prepare() factory method."""

    @pytest.fixture
    def sample_xlsx(self, tmp_path: Path) -> Path:
        """Create a sample workbook for testing."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        return path

    def test_inplace_edit_no_output(self, sample_xlsx: Path) -> None:
        """In-place edit when output_path is None."""
        session = EditSession.prepare(sample_xlsx, None)

        assert session.edit_path == sample_xlsx
        assert session.is_inplace is True
        assert session.input_path == sample_xlsx

    def test_inplace_edit_same_path(self, sample_xlsx: Path) -> None:
        """In-place edit when output equals input with force_inplace."""
        session = EditSession.prepare(sample_xlsx, sample_xlsx, force_inplace=True)

        assert session.edit_path == sample_xlsx
        assert session.is_inplace is True
        assert session.input_path == sample_xlsx

    def test_inplace_edit_default_behavior(self, sample_xlsx: Path) -> None:
        """In-place edit is default behavior when output same as input."""
        # With default force_inplace=True, this should work
        session = EditSession.prepare(sample_xlsx, sample_xlsx)

        assert session.edit_path == sample_xlsx
        assert session.is_inplace is True

    def test_copy_edit_creates_new_file(self, sample_xlsx: Path, tmp_path: Path) -> None:
        """Copy edit copies input to output before editing."""
        output_path = tmp_path / "output.xlsx"

        session = EditSession.prepare(sample_xlsx, output_path)

        # File should be copied
        assert session.edit_path == output_path
        assert session.is_inplace is False
        assert session.input_path == sample_xlsx
        assert output_path.exists()

        # Verify content was copied
        wb = Workbook()
        from openpyxl import load_workbook

        wb_loaded = load_workbook(str(output_path))
        assert wb_loaded.active["A1"].value == "Test"

    def test_copy_edit_creates_parent_dirs(self, sample_xlsx: Path, tmp_path: Path) -> None:
        """Copy edit creates parent directories when needed."""
        output_path = tmp_path / "nested" / "deep" / "output.xlsx"
        assert not output_path.parent.exists()

        session = EditSession.prepare(sample_xlsx, output_path)

        assert output_path.parent.exists()
        assert session.edit_path == output_path

    def test_copy_edit_fails_without_create_parents(
        self, sample_xlsx: Path, tmp_path: Path
    ) -> None:
        """Copy edit fails if parent doesn't exist and create_parents=False."""
        output_path = tmp_path / "nested" / "deep" / "output.xlsx"

        with pytest.raises(FileNotFoundError):
            EditSession.prepare(sample_xlsx, output_path, create_parents=False)

    def test_missing_input_file(self, tmp_path: Path) -> None:
        """Raises FileNotFoundError if input doesn't exist."""
        missing_path = tmp_path / "missing.xlsx"

        with pytest.raises(FileNotFoundError) as exc_info:
            EditSession.prepare(missing_path)

        assert "not found" in str(exc_info.value)

    def test_input_not_a_file(self, tmp_path: Path) -> None:
        """Raises error if input is a directory."""
        with pytest.raises(ValueError) as exc_info:
            EditSession.prepare(tmp_path)

        assert "not a file" in str(exc_info.value)


class TestEditSessionContextManager:
    """Tests for EditSession as context manager."""

    @pytest.fixture
    def sample_xlsx(self, tmp_path: Path) -> Path:
        """Create a sample workbook."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Original"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        return path

    def test_context_manager_enters_agent(self, sample_xlsx: Path) -> None:
        """Context manager enters and provides locked agent."""
        session = EditSession.prepare(sample_xlsx, None)

        with session as entered:
            assert entered is session
            assert session._agent is not None
            assert session.workbook is not None

    def test_context_manager_exits_and_saves(self, sample_xlsx: Path) -> None:
        """Context manager saves on successful exit."""
        session = EditSession.prepare(sample_xlsx, None)

        with session:
            session.workbook.active["A1"] = "Modified"

        # Verify save happened
        from openpyxl import load_workbook

        wb = load_workbook(str(sample_xlsx))
        assert wb.active["A1"].value == "Modified"

    def test_context_manager_no_save_on_exception(self, sample_xlsx: Path) -> None:
        """Context manager does NOT save on exception."""
        session = EditSession.prepare(sample_xlsx, None)

        try:
            with session:
                session.workbook.active["A1"] = "Modified"
                raise ValueError("Test exception")
        except ValueError:
            pass

        # Verify NO save happened
        from openpyxl import load_workbook

        wb = load_workbook(str(sample_xlsx))
        assert wb.active["A1"].value == "Original"

    def test_reentrant_raises_error(self, sample_xlsx: Path) -> None:
        """Entering twice raises RuntimeError."""
        session = EditSession.prepare(sample_xlsx, None)

        with session:
            with pytest.raises(RuntimeError, match="already entered"):
                session.__enter__()

    def test_access_outside_context_raises(self, sample_xlsx: Path) -> None:
        """Accessing workbook outside context raises RuntimeError."""
        session = EditSession.prepare(sample_xlsx, None)

        with pytest.raises(RuntimeError, match="outside context"):
            _ = session.workbook

    def test_agent_property_outside_context_raises(self, sample_xlsx: Path) -> None:
        """Accessing agent outside context raises RuntimeError."""
        session = EditSession.prepare(sample_xlsx, None)

        with pytest.raises(RuntimeError, match="outside context"):
            _ = session.agent


class TestEditSessionProperties:
    """Tests for EditSession properties."""

    @pytest.fixture
    def sample_xlsx(self, tmp_path: Path) -> Path:
        """Create a sample workbook."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        return path

    def test_version_hash_available(self, sample_xlsx: Path) -> None:
        """version_hash property returns geometry hash."""
        session = EditSession.prepare(sample_xlsx, None)

        with session:
            hash_val = session.version_hash
            assert isinstance(hash_val, str)
            assert hash_val.startswith("sha256:")

    def test_version_hash_outside_context_raises(self, sample_xlsx: Path) -> None:
        """version_hash outside context raises RuntimeError."""
        session = EditSession.prepare(sample_xlsx, None)

        with pytest.raises(RuntimeError, match="outside context"):
            _ = session.version_hash

    def test_file_hash_available(self, sample_xlsx: Path) -> None:
        """file_hash property returns file hash."""
        session = EditSession.prepare(sample_xlsx, None)

        with session:
            hash_val = session.file_hash
            assert isinstance(hash_val, str)
            assert hash_val.startswith("sha256:")

    def test_file_hash_outside_context_raises(self, sample_xlsx: Path) -> None:
        """file_hash outside context raises RuntimeError."""
        session = EditSession.prepare(sample_xlsx, None)

        with pytest.raises(RuntimeError, match="outside context"):
            _ = session.file_hash


class TestPrepareEditTarget:
    """Tests for prepare_edit_target() legacy compatibility function."""

    @pytest.fixture
    def sample_xlsx(self, tmp_path: Path) -> Path:
        """Create a sample workbook."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        return path

    def test_inplace_returns_same_path(self, sample_xlsx: Path) -> None:
        """In-place edit returns same path."""
        result = prepare_edit_target(sample_xlsx)
        assert result == sample_xlsx

    def test_copy_returns_output_path(self, sample_xlsx: Path, tmp_path: Path) -> None:
        """Copy edit returns output path and copies file."""
        output_path = tmp_path / "output.xlsx"

        result = prepare_edit_target(sample_xlsx, output_path)

        assert result == output_path
        assert output_path.exists()

    def test_string_paths_accepted(self, sample_xlsx: Path, tmp_path: Path) -> None:
        """Both Path and string inputs accepted."""
        output_path = tmp_path / "output.xlsx"

        result = prepare_edit_target(str(sample_xlsx), str(output_path))

        assert result == output_path


class TestEditSessionMacroHandling:
    """Tests for macro preservation through EditSession."""

    def test_xlsm_extension_preserves_vba(self, tmp_path: Path) -> None:
        """ExcelAgent auto-detects macros from .xlsm extension."""
        # Create a .xlsm file (just extension, no actual VBA for this test)
        wb = Workbook()
        path = tmp_path / "test.xlsm"
        wb.save(str(path))

        session = EditSession.prepare(path, None)

        with session as s:
            # Agent should have keep_vba=True based on extension
            assert s.agent._keep_vba is True

    def test_xlsx_extension_no_vba(self, tmp_path: Path) -> None:
        """ExcelAgent auto-detects no macros from .xlsx extension."""
        wb = Workbook()
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        session = EditSession.prepare(path, None)

        with session as s:
            # Agent should have keep_vba=False based on extension
            assert s.agent._keep_vba is False


class TestEditSessionConcurrency:
    """Tests for concurrent modification detection."""

    def test_concurrent_modification_detected(self, tmp_path: Path) -> None:
        """EditSession detects concurrent modifications via ExcelAgent."""
        from excel_agent.utils.exceptions import ConcurrentModificationError

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Original"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        session = EditSession.prepare(path, None)

        # Concurrent modification should be detected on save
        with pytest.raises(ConcurrentModificationError) as exc_info:
            with session:
                # Simulate external modification
                wb2 = Workbook()
                wb2.active["A1"] = "External"
                wb2.save(str(path))

                # This should raise ConcurrentModificationError when saving
                session.workbook.active["A1"] = "Internal"

        assert "modified by another process" in str(exc_info.value)


class TestEditSessionRealWorldPatterns:
    """Tests for common real-world usage patterns."""

    @pytest.fixture
    def sample_xlsx(self, tmp_path: Path) -> Path:
        """Create a sample workbook."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws["A2"] = "Data"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        return path

    def test_pattern_inplace_modification(self, sample_xlsx: Path) -> None:
        """Pattern: In-place modification of workbook."""
        session = EditSession.prepare(sample_xlsx, None)

        with session:
            ws = session.workbook.active
            ws["B1"] = "New Column"
            ws["B2"] = "New Data"

        # Verify modification persisted
        from openpyxl import load_workbook

        wb = load_workbook(str(sample_xlsx))
        assert wb.active["B1"].value == "New Column"
        assert wb.active["B2"].value == "New Data"

    def test_pattern_copy_before_modify(self, sample_xlsx: Path, tmp_path: Path) -> None:
        """Pattern: Copy before modify (clone semantics)."""
        output_path = tmp_path / "modified.xlsx"

        session = EditSession.prepare(sample_xlsx, output_path)

        with session:
            ws = session.workbook.active
            ws["A2"] = "Modified Data"

        # Original unchanged
        from openpyxl import load_workbook

        original = load_workbook(str(sample_xlsx))
        assert original.active["A2"].value == "Data"

        # Copy modified
        modified = load_workbook(str(output_path))
        assert modified.active["A2"].value == "Modified Data"

    def test_pattern_batch_operations(self, sample_xlsx: Path) -> None:
        """Pattern: Multiple operations in single session."""
        session = EditSession.prepare(sample_xlsx, None)

        with session:
            wb = session.workbook
            # Add sheet
            wb.create_sheet("NewSheet")
            # Modify data
            wb.active["A1"] = "Modified"
            # Get version hash for reporting
            version = session.version_hash

        # All changes persisted
        from openpyxl import load_workbook

        wb = load_workbook(str(sample_xlsx))
        assert "NewSheet" in wb.sheetnames
        assert wb.active["A1"].value == "Modified"

    def test_pattern_error_recovery(self, sample_xlsx: Path) -> None:
        """Pattern: Error doesn't corrupt original."""
        from openpyxl import load_workbook

        # Get original content from fixture
        original_wb = load_workbook(str(sample_xlsx))
        original_content = original_wb.active["A1"].value

        session = EditSession.prepare(sample_xlsx, None)

        try:
            with session:
                session.workbook.active["A1"] = "Modified"
                raise RuntimeError("Simulated error")
        except RuntimeError:
            pass

        # Original unchanged due to error
        wb = load_workbook(str(sample_xlsx))
        assert wb.active["A1"].value == original_content

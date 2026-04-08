"""
Shared test fixtures and configuration for excel-agent-tools.

Fixtures create temporary workbooks in isolated tmp_path directories,
ensuring zero test pollution. No fixture writes to the project directory.
"""

from __future__ import annotations

import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Workbook Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_workbook(tmp_path: Path) -> Path:
    """Create a basic 3-sheet workbook with data and formulas.

    Structure:
        Sheet1: Data + formulas (A1:C10 with SUM in C column)
        Sheet2: Cross-sheet reference (A1 = Sheet1!C10)
        Sheet3: Named range target

    Returns:
        Path to the created .xlsx file.
    """
    wb = Workbook()

    # --- Sheet1: Data + formulas ---
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Sheet1"
    ws1["A1"] = "Name"
    ws1["B1"] = "Value"
    ws1["C1"] = "Doubled"

    for i in range(2, 11):
        ws1[f"A{i}"] = f"Item {i - 1}"
        ws1[f"B{i}"] = (i - 1) * 10
        ws1[f"C{i}"] = f"=B{i}*2"

    ws1["B11"] = "=SUM(B2:B10)"
    ws1["C11"] = "=SUM(C2:C10)"

    # --- Sheet2: Cross-sheet reference ---
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Total from Sheet1"
    ws2["B1"] = "=Sheet1!B11"
    ws2["A2"] = "Double total"
    ws2["B2"] = "=B1*2"

    # --- Sheet3: Named range target ---
    ws3 = wb.create_sheet("Sheet3")
    ws3["A1"] = "Category"
    ws3["B1"] = "Amount"
    for i in range(2, 6):
        ws3[f"A{i}"] = f"Cat {i - 1}"
        ws3[f"B{i}"] = (i - 1) * 100

    # Define a named range
    from openpyxl.workbook.defined_name import DefinedName

    defn = DefinedName("SalesData", attr_text="Sheet3!$A$1:$B$5")
    wb.defined_names.add(defn)

    path = tmp_path / "sample.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def empty_workbook(tmp_path: Path) -> Path:
    """Create a minimal workbook with a single empty sheet.

    Returns:
        Path to the created .xlsx file.
    """
    wb = Workbook()
    path = tmp_path / "empty.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def formula_workbook(tmp_path: Path) -> Path:
    """Create a workbook with various formula patterns for dependency testing.

    Patterns:
        - Simple chain: A1 → B1 → C1
        - Cross-sheet: Sheet2!A1 → Sheet1!C1
        - Multi-reference: D1 = A1 + B1 + C1

    Returns:
        Path to the created .xlsx file.
    """
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Sheet1"

    ws1["A1"] = 10
    ws1["B1"] = "=A1*2"
    ws1["C1"] = "=B1+5"
    ws1["D1"] = "=A1+B1+C1"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "=Sheet1!C1"
    ws2["B1"] = "=A1*3"

    path = tmp_path / "formulas.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def circular_ref_workbook(tmp_path: Path) -> Path:
    """Create a workbook with intentional circular references.

    A1 = B1 + 1, B1 = A1 + 1 (circular)

    Returns:
        Path to the created .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "=B1+1"
    ws["B1"] = "=A1+1"

    path = tmp_path / "circular.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def large_workbook(tmp_path: Path) -> Path:
    """Create a 100k-row workbook for performance testing.

    Uses openpyxl write-only mode for speed. 100,000 rows × 5 columns.

    Returns:
        Path to the created .xlsx file.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Data")

    # Header
    ws.append(["ID", "Name", "Value", "Date", "Active"])

    # Data rows
    base_date = datetime.datetime(2026, 1, 1, tzinfo=datetime.UTC)
    for i in range(1, 100_001):
        ws.append(
            [
                i,
                f"Item {i}",
                i * 1.5,
                base_date + datetime.timedelta(days=i % 365),
                i % 2 == 0,
            ]
        )

    path = tmp_path / "large_dataset.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def styled_workbook(tmp_path: Path) -> Path:
    """Create a workbook with various cell styles for formatting tests.

    Returns:
        Path to the created .xlsx file.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    assert ws is not None

    ws["A1"] = "Bold Red"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FF0000")

    ws["B1"] = "Yellow Fill"
    ws["B1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws["C1"] = "Centered"
    ws["C1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["D1"] = 42.5
    ws["D1"].number_format = "$#,##0.00"

    ws["E1"] = "Bordered"
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ws["E1"].border = thin_border

    path = tmp_path / "styled.xlsx"
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# Utility Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def work_dir(tmp_path: Path) -> Path:
    """Create a temporary work directory (simulates /work/ for clone operations).

    Returns:
        Path to the work directory.
    """
    d = tmp_path / "work"
    d.mkdir()
    return d


@pytest.fixture
def output_dir(tmp_path: Path) -> Path:
    """Create a temporary output directory.

    Returns:
        Path to the output directory.
    """
    d = tmp_path / "output"
    d.mkdir()
    return d

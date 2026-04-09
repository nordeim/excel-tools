"""Tests for chunked I/O helper."""

from __future__ import annotations

import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_agent.core.chunked_io import (
    _serialize_cell_value,
    count_used_rows,
    read_range_chunked,
    read_range_full,
)


class TestSerializeCellValue:
    def test_none(self) -> None:
        assert _serialize_cell_value(None) is None

    def test_string(self) -> None:
        assert _serialize_cell_value("hello") == "hello"

    def test_number(self) -> None:
        assert _serialize_cell_value(42) == 42
        assert _serialize_cell_value(3.14) == 3.14

    def test_boolean(self) -> None:
        assert _serialize_cell_value(True) is True

    def test_datetime(self) -> None:
        dt = datetime.datetime(2026, 4, 8, 14, 30, tzinfo=datetime.timezone.utc)
        result = _serialize_cell_value(dt)
        assert "2026-04-08" in result

    def test_date(self) -> None:
        d = datetime.date(2026, 4, 8)
        assert _serialize_cell_value(d) == "2026-04-08"


class TestReadRangeFull:
    def test_basic_read(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Name"
        ws["B1"] = "Age"
        ws["A2"] = "Alice"
        ws["B2"] = 30
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        from openpyxl import load_workbook

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2 is not None
        result = read_range_full(ws2, 1, 1, 2, 2)
        assert len(result) == 2
        assert result[0] == ["Name", "Age"]
        assert result[1] == ["Alice", 30]


class TestReadRangeChunked:
    def test_chunked_read(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        for i in range(1, 101):
            ws.cell(row=i, column=1, value=i)
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        from openpyxl import load_workbook

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2 is not None

        chunks = list(read_range_chunked(ws2, 1, 1, 100, 1, chunk_size=30))
        assert len(chunks) == 4  # 30 + 30 + 30 + 10
        assert len(chunks[0]) == 30
        assert len(chunks[-1]) == 10

    def test_chunked_equals_full(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        for i in range(1, 51):
            ws.cell(row=i, column=1, value=f"row_{i}")
            ws.cell(row=i, column=2, value=i * 10)
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        from openpyxl import load_workbook

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2 is not None

        full = read_range_full(ws2, 1, 1, 50, 2)
        chunked_rows: list = []
        for chunk in read_range_chunked(ws2, 1, 1, 50, 2, chunk_size=20):
            chunked_rows.extend(chunk)
        assert full == chunked_rows


class TestCountUsedRows:
    def test_basic_count(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "data"
        ws["A5"] = "more"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        from openpyxl import load_workbook

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2 is not None
        assert count_used_rows(ws2) == 5

    def test_empty_sheet(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        assert count_used_rows(ws) == 0

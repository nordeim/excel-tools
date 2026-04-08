"""Tests for ApprovalTokenManager - HMAC-SHA256 scoped approval tokens."""

from __future__ import annotations

import time
from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_agent.governance.token_manager import (
    DEFAULT_TTL,
    MAX_TTL,
    ApprovalToken,
    ApprovalTokenManager,
    VALID_SCOPES,
)
from excel_agent.utils.exceptions import PermissionDeniedError


class TestApprovalTokenStructure:
    """Tests for ApprovalToken data structure."""

    def test_token_roundtrip(self) -> None:
        """Token serialization and deserialization should be reversible."""
        token = ApprovalToken(
            scope="sheet:delete",
            target_file_hash="sha256:abc123",
            nonce="def456",
            issued_at=1234567890.123456,
            ttl_seconds=300,
            signature="sig789",
        )

        token_str = token.to_string()
        restored = ApprovalToken.from_string(token_str)

        assert restored.scope == token.scope
        assert restored.target_file_hash == token.target_file_hash
        assert restored.nonce == token.nonce
        assert restored.issued_at == token.issued_at
        assert restored.ttl_seconds == token.ttl_seconds
        assert restored.signature == token.signature


class TestApprovalTokenManagerGenerate:
    """Tests for token generation."""

    def test_generate_valid_scope(self, tmp_path: Path) -> None:
        """Should generate token for valid scope."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        token_str = manager.generate_token(
            scope="sheet:delete",
            target_file=wb_path,
        )

        assert token_str
        assert len(token_str) > 0

    def test_generate_invalid_scope_raises(self, tmp_path: Path) -> None:
        """Should raise ValueError for invalid scope."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        with pytest.raises(ValueError, match="Invalid scope"):
            manager.generate_token(
                scope="invalid:scope",
                target_file=wb_path,
            )

    def test_generate_custom_ttl(self, tmp_path: Path) -> None:
        """Should accept custom TTL within range."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        token_str = manager.generate_token(
            scope="sheet:delete",
            target_file=wb_path,
            ttl_seconds=600,
        )

        token = ApprovalToken.from_string(token_str)
        assert token.ttl_seconds == 600

    def test_generate_ttl_too_high_raises(self, tmp_path: Path) -> None:
        """Should raise ValueError for TTL > MAX_TTL."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        with pytest.raises(ValueError, match="TTL must be"):
            manager.generate_token(
                scope="sheet:delete",
                target_file=wb_path,
                ttl_seconds=MAX_TTL + 1,
            )

    def test_generate_default_ttl(self, tmp_path: Path) -> None:
        """Should use DEFAULT_TTL when not specified."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        token_str = manager.generate_token(
            scope="sheet:delete",
            target_file=wb_path,
        )

        token = ApprovalToken.from_string(token_str)
        assert token.ttl_seconds == DEFAULT_TTL


class TestApprovalTokenManagerValidate:
    """Tests for token validation."""

    def test_validate_valid_token(self, tmp_path: Path) -> None:
        """Should validate a fresh token."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        token_str = manager.generate_token(
            scope="sheet:delete",
            target_file=wb_path,
        )

        # Should not raise
        manager.validate_token(
            token_string=token_str,
            scope="sheet:delete",
            file_path=wb_path,
        )

    def test_validate_missing_token_raises(self, tmp_path: Path) -> None:
        """Should raise PermissionDeniedError for missing token."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        with pytest.raises(PermissionDeniedError, match="No approval token"):
            manager.validate_token(
                token_string=None,
                scope="sheet:delete",
                file_path=wb_path,
            )

    def test_validate_wrong_scope_raises(self, tmp_path: Path) -> None:
        """Should raise PermissionDeniedError for scope mismatch."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        token_str = manager.generate_token(
            scope="sheet:delete",
            target_file=wb_path,
        )

        with pytest.raises(PermissionDeniedError, match="scope mismatch"):
            manager.validate_token(
                token_string=token_str,
                scope="range:delete",
                file_path=wb_path,
            )

    def test_validate_expired_token_raises(self, tmp_path: Path) -> None:
        """Should raise PermissionDeniedError for expired token."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        token_str = manager.generate_token(
            scope="sheet:delete",
            target_file=wb_path,
            ttl_seconds=1,  # Very short TTL
        )

        # Wait for expiration
        time.sleep(1.1)

        with pytest.raises(PermissionDeniedError, match="expired"):
            manager.validate_token(
                token_string=token_str,
                scope="sheet:delete",
                file_path=wb_path,
            )

    def test_validate_modified_file_raises(self, tmp_path: Path) -> None:
        """Should raise PermissionDeniedError if file was modified."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "original"
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        token_str = manager.generate_token(
            scope="sheet:delete",
            target_file=wb_path,
        )

        # Modify the file with different content
        wb2 = Workbook()
        ws2 = wb2.active
        assert ws2 is not None
        ws2["A1"] = "modified"  # Different content ensures different hash
        wb2.save(str(wb_path))

        with pytest.raises(PermissionDeniedError, match="file hash"):
            manager.validate_token(
                token_string=token_str,
                scope="sheet:delete",
                file_path=wb_path,
            )

    def test_validate_replay_raises(self, tmp_path: Path) -> None:
        """Should raise PermissionDeniedError for replay (already used)."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager = ApprovalTokenManager(secret="test-secret")
        token_str = manager.generate_token(
            scope="sheet:delete",
            target_file=wb_path,
        )

        # First use succeeds
        manager.validate_token(
            token_string=token_str,
            scope="sheet:delete",
            file_path=wb_path,
        )

        # Second use (replay) fails - use same manager to preserve nonce tracking
        with pytest.raises(PermissionDeniedError, match="already used"):
            manager.validate_token(
                token_string=token_str,
                scope="sheet:delete",
                file_path=wb_path,
            )


class TestApprovalTokenManagerSignature:
    """Tests for HMAC signature verification."""

    def test_different_secrets_fail(self, tmp_path: Path) -> None:
        """Tokens from different secrets should not validate."""
        wb = Workbook()
        wb_path = tmp_path / "test.xlsx"
        wb.save(str(wb_path))

        manager1 = ApprovalTokenManager(secret="secret-1")
        manager2 = ApprovalTokenManager(secret="secret-2")

        token_str = manager1.generate_token(
            scope="sheet:delete",
            target_file=wb_path,
        )

        with pytest.raises(PermissionDeniedError, match="Invalid token"):
            manager2.validate_token(
                token_string=token_str,
                scope="sheet:delete",
                file_path=wb_path,
            )


class TestValidScopes:
    """Tests for valid scope definitions."""

    def test_expected_scopes_present(self) -> None:
        """Should have expected scopes defined."""
        expected = {
            "sheet:delete",
            "sheet:rename",
            "range:delete",
            "formula:convert",
            "macro:remove",
            "macro:inject",
            "structure:modify",
        }
        assert expected.issubset(VALID_SCOPES)

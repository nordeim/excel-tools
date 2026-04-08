"""Tests for DependencyTracker - the formula dependency graph engine."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_agent.core.dependency import DependencyTracker, ImpactReport


class TestDependencyTrackerBasic:
    """Basic dependency tracker functionality tests."""

    def test_empty_workbook_empty_graph(self, tmp_path: Path) -> None:
        """Empty workbook should produce empty graph."""
        wb = Workbook()
        wb_path = tmp_path / "empty.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        assert tracker.is_built
        stats = tracker.get_stats()
        assert stats["total_formulas"] == 0
        assert stats["total_edges"] == 0
        assert stats["circular_chains"] == 0

    def test_value_cells_not_in_graph(self, tmp_path: Path) -> None:
        """Cells with values (not formulas) should not be in graph."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = 10
        ws["B1"] = "hello"

        wb_path = tmp_path / "values.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        stats = tracker.get_stats()
        assert stats["total_formulas"] == 0

    def test_single_dependency(self, tmp_path: Path) -> None:
        """A1=B1 should create edge: B1 -> A1."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=B1"

        wb_path = tmp_path / "single.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        # A1 depends on B1
        precedents = tracker.find_precedents("A1")
        assert "Sheet!B1" in precedents or any("B1" in p for p in precedents)

        # B1 is depended on by A1
        dependents = tracker.find_dependents("B1")
        assert any("A1" in d for d in dependents)


class TestDependencyTrackerChain:
    """Tests for dependency chains."""

    def test_chain_a1_b1_c1(self, tmp_path: Path) -> None:
        """Chain A1=B1, B1=C1: deleting C1 affects B1 and A1."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=B1"
        ws["B1"] = "=C1"
        ws["C1"] = 5

        wb_path = tmp_path / "chain.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        # C1's dependents should include B1 and A1
        dependents = tracker.find_dependents("C1")
        dependent_coords = {d.split("!")[-1] for d in dependents}
        assert "B1" in dependent_coords
        assert "A1" in dependent_coords


class TestDependencyTrackerCrossSheet:
    """Tests for cross-sheet dependencies."""

    def test_cross_sheet_dependency(self, tmp_path: Path) -> None:
        """Sheet2!A1 = Sheet1!B1 should be tracked."""
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "Sheet1"
        ws1["B1"] = 10

        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "=Sheet1!B1"

        wb_path = tmp_path / "cross.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        # Sheet1!B1's dependents should include Sheet2!A1
        dependents = tracker.find_dependents("Sheet1!B1")
        assert any("Sheet2!A1" in d for d in dependents)


class TestDependencyTrackerCircular:
    """Tests for circular reference detection."""

    def test_two_cell_circular(self, tmp_path: Path) -> None:
        """A1=B1, B1=A1 should be detected as circular."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=B1+1"
        ws["B1"] = "=A1+1"

        wb_path = tmp_path / "circular.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        cycles = tracker.detect_circular_references()
        assert len(cycles) > 0
        # Check that A1 and B1 are in a cycle together
        cycle_nodes = set()
        for cycle in cycles:
            cycle_nodes.update(cycle)
        assert any("A1" in n for n in cycle_nodes)
        assert any("B1" in n for n in cycle_nodes)

    def test_three_cell_circular(self, tmp_path: Path) -> None:
        """A1=B1, B1=C1, C1=A1 should be detected as circular."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=B1"
        ws["B1"] = "=C1"
        ws["C1"] = "=A1"

        wb_path = tmp_path / "circular3.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        cycles = tracker.detect_circular_references()
        assert len(cycles) > 0


class TestDependencyTrackerImpactReport:
    """Tests for impact report generation."""

    def test_safe_operation_no_dependencies(self, tmp_path: Path) -> None:
        """Deleting a cell with no dependents should report 'safe'."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = 10  # Value, not formula
        ws["B1"] = "=A1"  # Formula depends on A1

        wb_path = tmp_path / "safe.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        # Check impact of modifying C1 (no dependents)
        report = tracker.impact_report("C1")
        assert report.status == "safe"
        assert report.broken_references == 0

    def test_warning_operation_few_dependencies(self, tmp_path: Path) -> None:
        """Deleting a cell with <10 dependents should report 'warning'."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = 10

        # Create 5 formulas that depend on A1
        for i in range(1, 6):
            ws[f"B{i}"] = f"=A1*{i}"

        wb_path = tmp_path / "warning.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        report = tracker.impact_report("A1")
        assert report.status in ("warning", "critical")
        assert report.broken_references == 5

    def test_critical_operation_many_dependencies(self, tmp_path: Path) -> None:
        """Deleting a cell with >10 dependents should report 'critical'."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = 10

        # Create 15 formulas that depend on A1
        for i in range(1, 16):
            ws[f"B{i}"] = f"=A1*{i}"

        wb_path = tmp_path / "critical.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        report = tracker.impact_report("A1")
        assert report.status == "critical"
        assert report.broken_references == 15
        assert len(report.sample_errors) <= 10  # First 10 only

    def test_impact_report_suggestion_present(self, tmp_path: Path) -> None:
        """Impact report should include prescriptive suggestion."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = 10
        ws["B1"] = "=A1"

        wb_path = tmp_path / "suggestion.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        report = tracker.impact_report("A1")
        assert report.suggestion
        assert len(report.suggestion) > 0


class TestDependencyTrackerStats:
    """Tests for dependency graph statistics."""

    def test_stats_accuracy(self, tmp_path: Path) -> None:
        """Stats should accurately reflect graph contents."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=B1+C1"

        wb_path = tmp_path / "stats.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        stats = tracker.get_stats()
        assert stats["total_formulas"] == 1
        assert stats["total_edges"] == 2  # A1 depends on B1 and C1


class TestDependencyTrackerAdjacencyList:
    """Tests for adjacency list export."""

    def test_adjacency_list_export(self, tmp_path: Path) -> None:
        """Adjacency list should be JSON-serializable."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=B1"

        wb_path = tmp_path / "adj.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        tracker.build_graph()

        adj_list = tracker.get_adjacency_list()
        assert isinstance(adj_list, dict)
        # Should be JSON-serializable
        import json

        json.dumps(adj_list)


class TestDependencyTrackerNotBuiltError:
    """Tests for error when methods called before build_graph."""

    def test_find_dependents_before_build_raises(self, tmp_path: Path) -> None:
        """find_dependents should raise if graph not built."""
        wb = Workbook()
        wb_path = tmp_path / "notbuilt.xlsx"
        wb.save(str(wb_path))

        from openpyxl import load_workbook

        tracker = DependencyTracker(load_workbook(str(wb_path)))
        # Don't build graph

        with pytest.raises(RuntimeError, match="not built"):
            tracker.find_dependents("A1")

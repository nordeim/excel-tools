"""Tests for merge/unmerge cell operations."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class TestMergePreCheck:
    """Tests for detecting hidden data before merge."""

    def test_empty_range_no_warning(self, tmp_path: Path) -> None:
        """Merging a range where non-anchor cells are empty should succeed."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Title"
        # B1, C1 are empty — should be safe to merge
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2 is not None
        # Check non-anchor cells
        has_data = False
        for col in range(2, 4):  # B1, C1
            if ws2.cell(row=1, column=col).value is not None:
                has_data = True
        assert not has_data

    def test_data_in_non_anchor_detected(self, tmp_path: Path) -> None:
        """Non-anchor cells with data should be detected."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Title"
        ws["B1"] = "Hidden Data"
        ws["C1"] = 42
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2 is not None
        hidden = []
        for col in range(2, 4):
            cell = ws2.cell(row=1, column=col)
            if cell.value is not None:
                hidden.append(f"{get_column_letter(col)}1")
        assert len(hidden) == 2
        assert "B1" in hidden
        assert "C1" in hidden


class TestMergeUnmerge:
    """Tests for openpyxl merge/unmerge behavior."""

    def test_merge_and_unmerge_roundtrip(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Merged Title"
        ws.merge_cells("A1:D1")
        path = tmp_path / "merged.xlsx"
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2 is not None
        assert len(list(ws2.merged_cells.ranges)) == 1

        ws2.unmerge_cells("A1:D1")
        assert len(list(ws2.merged_cells.ranges)) == 0

    def test_merge_removes_non_anchor_values(self, tmp_path: Path) -> None:
        """After merge, non-anchor cells have value=None."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Keep"
        ws["B1"] = "Lost"
        ws["C1"] = "Lost"
        ws.merge_cells("A1:C1")
        path = tmp_path / "merged.xlsx"
        wb.save(str(path))

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2 is not None
        assert ws2["A1"].value == "Keep"
        # Non-anchor cells should be None after merge
        assert ws2["B1"].value is None
        assert ws2["C1"].value is None

    def test_unmerge_all(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.merge_cells("A1:C1")
        ws.merge_cells("A3:B5")
        assert len(list(ws.merged_cells.ranges)) == 2

        # Unmerge all
        for mr in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mr))
        assert len(list(ws.merged_cells.ranges)) == 0

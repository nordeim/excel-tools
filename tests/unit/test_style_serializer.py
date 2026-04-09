"""Tests for cell style serialization."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from excel_agent.core.style_serializer import (
    serialize_alignment,
    serialize_border,
    serialize_cell_style,
    serialize_fill,
    serialize_font,
)


class TestSerializeFont:
    def test_basic_font(self) -> None:
        font = Font(name="Arial", size=12, bold=True, color="FF0000")
        result = serialize_font(font)
        assert result["name"] == "Arial"
        assert result["size"] == 12
        assert result["bold"] is True
        assert result["color"] is not None

    def test_default_font(self) -> None:
        font = Font()
        result = serialize_font(font)
        # Default Font has None for name, so it won't be in result
        assert "name" not in result or result.get("name") is None


class TestSerializeFill:
    def test_solid_fill(self) -> None:
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        result = serialize_fill(fill)
        assert result["patternType"] == "solid"

    def test_no_fill(self) -> None:
        fill = PatternFill()
        result = serialize_fill(fill)
        assert "patternType" not in result or result.get("patternType") is None


class TestSerializeBorder:
    def test_with_borders(self) -> None:
        border = Border(
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thick"),
        )
        result = serialize_border(border)
        assert result["top"]["style"] == "thin"
        assert result["bottom"]["style"] == "thick"

    def test_no_borders(self) -> None:
        border = Border()
        result = serialize_border(border)
        assert result == {}


class TestSerializeAlignment:
    def test_centered(self) -> None:
        alignment = Alignment(horizontal="center", vertical="center")
        result = serialize_alignment(alignment)
        assert result["horizontal"] == "center"
        assert result["vertical"] == "center"


class TestSerializeCellStyle:
    def test_styled_cell(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Styled"
        ws["A1"].font = Font(bold=True, size=14, color="FF0000")
        ws["A1"].fill = PatternFill("solid", fgColor="FFFF00")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].number_format = "$#,##0.00"

        result = serialize_cell_style(ws["A1"])
        assert result["font"]["bold"] is True
        assert result["fill"]["patternType"] == "solid"
        assert result["alignment"]["horizontal"] == "center"
        assert result["number_format"] == "$#,##0.00"

"""Tests for ExcelAgent context manager."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_agent.core.agent import ExcelAgent
from excel_agent.utils.exceptions import (
    ConcurrentModificationError,
    ExcelFileNotFoundError,
    ValidationError,
)


class TestExcelAgentBasic:
    """Basic ExcelAgent functionality tests."""

    def test_successful_enter_exit(self, tmp_path: Path) -> None:
        """Test successful load, modify, save cycle."""
        # Create a workbook
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active["A1"] = "Original"
        wb.save(str(wb_path))

        # Modify via ExcelAgent
        with ExcelAgent(wb_path, mode="rw") as agent:
            agent.workbook.active["A1"] = "Modified"

        # Verify file was saved
        wb2 = Workbook()
        wb2 = Workbook()
        from openpyxl import load_workbook

        wb2 = load_workbook(str(wb_path))
        assert wb2.active["A1"].value == "Modified"

    def test_read_only_mode_no_save(self, tmp_path: Path) -> None:
        """Read-only mode should not save changes."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active["A1"] = "Original"
        wb.save(str(wb_path))

        # Try to modify in read-only mode
        with ExcelAgent(wb_path, mode="r") as agent:
            agent.workbook.active["A1"] = "Modified"

        # Verify file was NOT saved
        from openpyxl import load_workbook

        wb2 = load_workbook(str(wb_path))
        assert wb2.active["A1"].value == "Original"

    def test_file_not_found_raises_error(self, tmp_path: Path) -> None:
        """Non-existent file should raise ExcelFileNotFoundError."""
        nonexistent = tmp_path / "does_not_exist.xlsx"
        with pytest.raises(ExcelFileNotFoundError), ExcelAgent(nonexistent, mode="rw"):
            pass

    def test_invalid_mode_raises_validation_error(self, tmp_path: Path) -> None:
        """Invalid mode should raise ValidationError."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))

        with pytest.raises(ValidationError, match="Invalid mode"):
            ExcelAgent(wb_path, mode="invalid")

    def test_workbook_access_outside_context_raises(self, tmp_path: Path) -> None:
        """Accessing workbook outside context should raise RuntimeError."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))

        agent = ExcelAgent(wb_path, mode="rw")
        with pytest.raises(RuntimeError, match="outside context"):
            _ = agent.workbook


class TestExcelAgentHash:
    """Tests for hash-related functionality."""

    def test_version_hash_returns_sha256_format(self, tmp_path: Path) -> None:
        """version_hash should return 'sha256:...' format."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))

        with ExcelAgent(wb_path, mode="rw") as agent:
            h = agent.version_hash
            assert h.startswith("sha256:")
            assert len(h) == 7 + 64

    def test_file_hash_returns_sha256_format(self, tmp_path: Path) -> None:
        """file_hash should return 'sha256:...' format."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))

        with ExcelAgent(wb_path, mode="rw") as agent:
            h = agent.file_hash
            assert h.startswith("sha256:")
            assert len(h) == 7 + 64


class TestExcelAgentConcurrentModification:
    """Tests for concurrent modification detection."""

    def test_concurrent_modification_detected(self, tmp_path: Path) -> None:
        """Should detect external modification during session."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active["A1"] = "Original"
        wb.save(str(wb_path))

        # Enter context but don't exit yet
        agent = ExcelAgent(wb_path, mode="rw")
        agent.__enter__()

        # Simulate external modification
        wb2 = Workbook()
        from openpyxl import load_workbook

        wb2 = load_workbook(str(wb_path))
        wb2.active["A1"] = "Externally Modified"
        wb2.save(str(wb_path))

        # Exit should detect modification and refuse to save
        with pytest.raises(ConcurrentModificationError):
            agent.__exit__(None, None, None)

    def test_verify_no_concurrent_modification_raises_on_change(self, tmp_path: Path) -> None:
        """verify_no_concurrent_modification should raise on external change."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))

        # Enter context manually to control the flow
        agent = ExcelAgent(wb_path, mode="rw")
        agent.__enter__()

        try:
            # Simulate external modification
            from openpyxl import load_workbook

            wb2 = load_workbook(str(wb_path))
            wb2.active["A1"] = "Changed"
            wb2.save(str(wb_path))

            # Should raise when we try to verify
            with pytest.raises(ConcurrentModificationError):
                agent.verify_no_concurrent_modification()
        finally:
            # Clean up: suppress the concurrent modification error during exit
            # since we've already verified it works
            try:
                agent.__exit__(None, None, None)
            except ConcurrentModificationError:
                pass  # Expected - we've already tested the verification


class TestExcelAgentException:
    """Tests for exception handling."""

    def test_exception_releases_lock_no_save(self, tmp_path: Path) -> None:
        """Exception in context body should release lock and not save."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active["A1"] = "Original"
        wb.save(str(wb_path))

        with pytest.raises(ValueError, match="deliberate error"):
            with ExcelAgent(wb_path, mode="rw") as agent:
                agent.workbook.active["A1"] = "Modified"
                raise ValueError("deliberate error")

        # Verify file was NOT saved
        from openpyxl import load_workbook

        wb2 = load_workbook(str(wb_path))
        assert wb2.active["A1"].value == "Original"

    def test_lock_released_after_exception(self, tmp_path: Path) -> None:
        """Lock should be released after exception."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))

        from excel_agent.core.locking import FileLock

        try:
            with ExcelAgent(wb_path, mode="rw"):
                raise ValueError("test")
        except ValueError:
            pass

        # Lock should be released - we can acquire it again
        assert not FileLock.is_locked(wb_path)


class TestExcelAgentVBA:
    """Tests for VBA/xlsm handling."""

    def test_xlsm_auto_detects_keep_vba(self, tmp_path: Path) -> None:
        """.xlsm extension should auto-detect keep_vba=True."""
        # Note: We can't easily create a real .xlsm without a VBA project,
        # but we can test that the auto-detection logic works by checking
        # the internal state
        wb_path = tmp_path / "test.xlsm"
        wb = Workbook()
        wb.save(str(wb_path))

        agent = ExcelAgent(wb_path, mode="rw")
        assert agent._keep_vba is True

    def test_xlsx_defaults_keep_vba_false(self, tmp_path: Path) -> None:
        """.xlsx extension should default keep_vba=False."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))

        agent = ExcelAgent(wb_path, mode="rw")
        assert agent._keep_vba is False

    def test_explicit_keep_vba_overrides_auto_detection(self, tmp_path: Path) -> None:
        """Explicit keep_vba should override auto-detection."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))

        agent = ExcelAgent(wb_path, mode="rw", keep_vba=True)
        assert agent._keep_vba is True


class TestExcelAgentProperties:
    """Tests for ExcelAgent properties."""

    def test_path_property(self, tmp_path: Path) -> None:
        """path property should return resolved Path."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))

        agent = ExcelAgent(wb_path, mode="rw")
        assert agent.path == wb_path.resolve()

    def test_lock_timeout_respected(self, tmp_path: Path) -> None:
        """Custom lock timeout should be respected."""
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))

        agent = ExcelAgent(wb_path, mode="rw", lock_timeout=60.0)
        assert agent._lock_timeout == 60.0

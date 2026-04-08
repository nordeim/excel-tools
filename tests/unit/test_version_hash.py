"""Tests for geometry-aware workbook hashing."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from excel_agent.core.version_hash import (
    compute_file_hash,
    compute_sheet_hash,
    compute_workbook_hash,
)


class TestWorkbookHash:
    """Tests for compute_workbook_hash."""

    def test_identical_workbooks_same_hash(self, tmp_path: Path) -> None:
        """Two structurally identical workbooks should produce the same hash."""
        wb1 = Workbook()
        ws1 = wb1.active
        assert ws1 is not None
        ws1["A1"] = "=B1+1"
        ws1["B1"] = 10
        path1 = tmp_path / "wb1.xlsx"
        wb1.save(str(path1))

        wb2 = Workbook()
        ws2 = wb2.active
        assert ws2 is not None
        ws2["A1"] = "=B1+1"
        ws2["B1"] = 10
        path2 = tmp_path / "wb2.xlsx"
        wb2.save(str(path2))

        from openpyxl import load_workbook as lw

        h1 = compute_workbook_hash(lw(str(path1)))
        h2 = compute_workbook_hash(lw(str(path2)))
        assert h1 == h2

    def test_value_change_same_hash(self, tmp_path: Path) -> None:
        """Changing a cell VALUE should NOT change the geometry hash."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=B1+1"
        ws["B1"] = 10

        h1 = compute_workbook_hash(wb)

        ws["B1"] = 9999  # Value change only

        h2 = compute_workbook_hash(wb)
        assert h1 == h2

    def test_formula_change_different_hash(self, tmp_path: Path) -> None:
        """Changing a cell FORMULA should change the geometry hash."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=B1+1"

        h1 = compute_workbook_hash(wb)

        ws["A1"] = "=B1+2"

        h2 = compute_workbook_hash(wb)
        assert h1 != h2

    def test_sheet_rename_different_hash(self) -> None:
        """Renaming a sheet should change the hash."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Data"

        h1 = compute_workbook_hash(wb)

        ws.title = "NewName"

        h2 = compute_workbook_hash(wb)
        assert h1 != h2

    def test_sheet_add_different_hash(self) -> None:
        """Adding a sheet should change the hash."""
        wb = Workbook()
        h1 = compute_workbook_hash(wb)

        wb.create_sheet("NewSheet")

        h2 = compute_workbook_hash(wb)
        assert h1 != h2

    def test_sheet_remove_different_hash(self) -> None:
        """Removing a sheet should change the hash."""
        wb = Workbook()
        wb.create_sheet("Extra")
        h1 = compute_workbook_hash(wb)

        wb.remove(wb["Extra"])

        h2 = compute_workbook_hash(wb)
        assert h1 != h2

    def test_sheet_reorder_different_hash(self) -> None:
        """Reordering sheets should change the hash."""
        wb = Workbook()
        wb.create_sheet("Alpha")
        wb.create_sheet("Beta")
        h1 = compute_workbook_hash(wb)

        wb.move_sheet("Beta", offset=-1)

        h2 = compute_workbook_hash(wb)
        assert h1 != h2

    def test_hash_format(self) -> None:
        """Hash should be in 'sha256:...' format."""
        wb = Workbook()
        h = compute_workbook_hash(wb)
        assert h.startswith("sha256:")
        # SHA256 hex digest is 64 characters
        assert len(h) == 7 + 64  # "sha256:" + 64 hex chars

    def test_empty_workbook_produces_hash(self) -> None:
        """Even an empty workbook should produce a valid hash."""
        wb = Workbook()
        h = compute_workbook_hash(wb)
        assert h.startswith("sha256:")


class TestSheetHash:
    """Tests for compute_sheet_hash."""

    def test_returns_valid_hash(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        h = compute_sheet_hash(ws)
        assert h.startswith("sha256:")


class TestFileHash:
    """Tests for compute_file_hash."""

    def test_file_hash_changes_on_modification(self, tmp_path: Path) -> None:
        f = tmp_path / "test.txt"
        f.write_text("hello")
        h1 = compute_file_hash(f)

        f.write_text("world")
        h2 = compute_file_hash(f)

        assert h1 != h2

    def test_same_content_same_hash(self, tmp_path: Path) -> None:
        f1 = tmp_path / "a.txt"
        f2 = tmp_path / "b.txt"
        f1.write_text("identical content")
        f2.write_text("identical content")

        assert compute_file_hash(f1) == compute_file_hash(f2)

    def test_hash_format(self, tmp_path: Path) -> None:
        f = tmp_path / "test.txt"
        f.write_text("data")
        h = compute_file_hash(f)
        assert h.startswith("sha256:")
        assert len(h) == 7 + 64

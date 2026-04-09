"""Integration tests for cell operation tools via subprocess."""

from __future__ import annotations

import json
import shutil
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def _run_tool(tool_module: str, *args: str) -> tuple[dict, int]:
    """Run a CLI tool and return (parsed_json, return_code)."""
    result = subprocess.run(
        [sys.executable, "-m", f"excel_agent.tools.{tool_module}", *args],
        capture_output=True,
        text=True,
        timeout=30,
    )
    data = json.loads(result.stdout) if result.stdout.strip() else {}
    return data, result.returncode


@pytest.fixture
def merge_workbook(tmp_path: Path) -> Path:
    """Create a workbook suitable for merge/unmerge testing."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws["A1"] = "Title"
    ws["B1"] = "Hidden"
    ws["C1"] = "Also Hidden"
    ws["A3"] = "Data Row"
    ws["B3"] = 100
    ws["C3"] = 200
    path = tmp_path / "merge_test.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def ref_workbook(tmp_path: Path) -> Path:
    """Create a workbook suitable for reference update testing."""
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Sheet1"
    ws1["A1"] = 10
    ws1["A2"] = 20
    ws1["B1"] = "=A1*2"
    ws1["B2"] = "=A2*3"
    ws1["C1"] = "=B1+B2"
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "=Sheet1!C1"
    path = tmp_path / "ref_test.xlsx"
    wb.save(str(path))
    return path


class TestMergeCells:
    def test_merge_empty_range(self, merge_workbook: Path, tmp_path: Path) -> None:
        work = tmp_path / "work.xlsx"
        shutil.copy2(merge_workbook, work)

        # Merge A3:C3 where B3 and C3 have data — should warn without --force
        data, code = _run_tool(
            "cells.xls_merge_cells",
            "--input",
            str(work),
            "--output",
            str(work),
            "--sheet",
            "Sheet1",
            "--range",
            "A3:C3",
        )
        # Should return warning because B3 and C3 have data
        assert code == 1 or data.get("status") == "warning"

    def test_merge_with_force(self, merge_workbook: Path, tmp_path: Path) -> None:
        work = tmp_path / "work.xlsx"
        shutil.copy2(merge_workbook, work)

        data, code = _run_tool(
            "cells.xls_merge_cells",
            "--input",
            str(work),
            "--output",
            str(work),
            "--sheet",
            "Sheet1",
            "--range",
            "A3:C3",
            "--force",
        )
        assert code == 0
        wb = load_workbook(str(work))
        ws = wb["Sheet1"]
        assert len(list(ws.merged_cells.ranges)) == 1


class TestUnmergeCells:
    def test_unmerge_specific(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws["A1"] = "Merged"
        ws.merge_cells("A1:C1")
        path = tmp_path / "merged.xlsx"
        wb.save(str(path))

        data, code = _run_tool(
            "cells.xls_unmerge_cells",
            "--input",
            str(path),
            "--output",
            str(path),
            "--sheet",
            "Sheet1",
            "--range",
            "A1:C1",
        )
        assert code == 0
        assert data["data"]["count"] == 1

        wb2 = load_workbook(str(path))
        assert len(list(wb2["Sheet1"].merged_cells.ranges)) == 0

    def test_unmerge_all(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws.merge_cells("A1:C1")
        ws.merge_cells("A3:B5")
        path = tmp_path / "multi_merged.xlsx"
        wb.save(str(path))

        data, code = _run_tool(
            "cells.xls_unmerge_cells",
            "--input",
            str(path),
            "--output",
            str(path),
            "--sheet",
            "Sheet1",
        )
        assert code == 0
        assert data["data"]["count"] == 2


class TestDeleteRange:
    def test_without_token_fails(self, ref_workbook: Path, tmp_path: Path) -> None:
        work = tmp_path / "work.xlsx"
        shutil.copy2(ref_workbook, work)

        data, code = _run_tool(
            "cells.xls_delete_range",
            "--input",
            str(work),
            "--output",
            str(work),
            "--sheet",
            "Sheet1",
            "--range",
            "A1:A2",
            "--shift",
            "up",
        )
        assert code == 1  # Missing token


class TestUpdateReferences:
    def test_basic_update(self, ref_workbook: Path, tmp_path: Path) -> None:
        work = tmp_path / "work.xlsx"
        shutil.copy2(ref_workbook, work)

        data, code = _run_tool(
            "cells.xls_update_references",
            "--input",
            str(work),
            "--output",
            str(work),
            "--updates",
            '[{"old": "Sheet1!A1", "new": "Sheet1!A5"}]',
        )
        assert code == 0
        assert data["data"]["formulas_updated"] >= 1

        wb = load_workbook(str(work))
        # B1 was =A1*2, should now reference A5
        formula = wb["Sheet1"]["B1"].value
        assert "A5" in formula

    def test_cross_sheet_update(self, ref_workbook: Path, tmp_path: Path) -> None:
        work = tmp_path / "work.xlsx"
        shutil.copy2(ref_workbook, work)

        data, code = _run_tool(
            "cells.xls_update_references",
            "--input",
            str(work),
            "--output",
            str(work),
            "--updates",
            '[{"old": "Sheet1!C1", "new": "Sheet1!C10"}]',
        )
        assert code == 0

        wb = load_workbook(str(work))
        # Sheet2!A1 was =Sheet1!C1, should now reference C10
        formula = wb["Sheet2"]["A1"].value
        assert "C10" in formula

    def test_no_matches_zero_updates(self, ref_workbook: Path, tmp_path: Path) -> None:
        work = tmp_path / "work.xlsx"
        shutil.copy2(ref_workbook, work)

        data, code = _run_tool(
            "cells.xls_update_references",
            "--input",
            str(work),
            "--output",
            str(work),
            "--updates",
            '[{"old": "Sheet1!Z99", "new": "Sheet1!Z100"}]',
        )
        assert code == 0
        assert data["data"]["formulas_updated"] == 0

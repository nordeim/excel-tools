"""Integration tests for export workflow.

Tests end-to-end export workflows combining multiple export operations.
"""

from __future__ import annotations

import csv
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def run_tool(tool: str, *args: str) -> tuple[dict, int]:
    """Helper to run an export tool and parse output."""
    cmd = [
        sys.executable,
        "-m",
        f"excel_agent.tools.export.{tool}",
        *args,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    try:
        output = json.loads(result.stdout) if result.stdout else {}
    except json.JSONDecodeError:
        output = {"parse_error": True, "stdout": result.stdout, "stderr": result.stderr}
    return output, result.returncode


@pytest.fixture
def data_workbook(tmp_path: Path) -> Path:
    """Create a workbook with data for export tests."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sales"

    # Headers
    ws["A1"] = "Product"
    ws["B1"] = "Q1"
    ws["C1"] = "Q2"
    ws["D1"] = "Total"

    # Data rows
    for i in range(2, 11):
        ws[f"A{i}"] = f"Product {i - 1}"
        ws[f"B{i}"] = i * 100
        ws[f"C{i}"] = i * 150
        ws[f"D{i}"] = f"=B{i}+C{i}"  # Formula

    path = tmp_path / "data.xlsx"
    wb.save(str(path))
    return path


class TestFullExportWorkflow:
    """Test complete export workflow."""

    def test_export_to_all_formats(self, data_workbook: Path, tmp_path: Path):
        """Test exporting workbook to CSV, JSON, and PDF."""
        # Step 1: Export to CSV
        csv_path = tmp_path / "output.csv"
        csv_out, csv_exit = run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(csv_path),
        )
        assert csv_exit == 0, f"CSV export failed: {csv_out}"
        assert csv_path.exists()

        # Step 2: Export to JSON
        json_path = tmp_path / "output.json"
        json_out, json_exit = run_tool(
            "xls_export_json",
            "--input",
            str(data_workbook),
            "--outfile",
            str(json_path),
            "--orient",
            "records",
        )
        assert json_exit == 0, f"JSON export failed: {json_out}"
        assert json_path.exists()

        # Step 3: Export to PDF (may fail if LibreOffice not installed)
        pdf_path = tmp_path / "output.pdf"
        pdf_out, pdf_exit = run_tool(
            "xls_export_pdf",
            "--input",
            str(data_workbook),
            "--outfile",
            str(pdf_path),
            "--timeout",
            "30",
        )
        # PDF may fail if LibreOffice not available, that's OK
        if pdf_exit == 0:
            assert pdf_path.exists()
            assert pdf_path.stat().st_size > 0

    def test_csv_roundtrip(self, data_workbook: Path, tmp_path: Path):
        """Test round-trip: Export CSV → Import to new workbook → Compare."""
        # Export to CSV
        csv_path = tmp_path / "exported.csv"
        run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(csv_path),
        )

        # Import back
        new_wb = Workbook()
        new_ws = new_wb.active
        assert new_ws is not None

        with open(csv_path) as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    new_ws.cell(row=row_idx, column=col_idx, value=value)

        # Verify basic data preserved
        assert new_ws["A1"].value == "Product"
        assert "Product 1" in str(new_ws["A2"].value)

    def test_json_api_simulation(self, data_workbook: Path, tmp_path: Path):
        """Test JSON export for API consumption."""
        json_path = tmp_path / "api_data.json"
        output, exit_code = run_tool(
            "xls_export_json",
            "--input",
            str(data_workbook),
            "--outfile",
            str(json_path),
            "--orient",
            "records",
        )

        assert exit_code == 0

        # Load and verify JSON structure (typical API payload)
        with open(json_path) as f:
            data = json.load(f)

        # Simulate API response validation
        assert isinstance(data, list)
        for record in data:
            assert "Product" in record
            assert "Q1" in record
            assert "Q2" in record

    def test_multi_sheet_export(self, data_workbook: Path, tmp_path: Path):
        """Test exporting different sheets."""
        # Add second sheet
        wb = load_workbook(str(data_workbook))
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Total Sales"
        ws2["B1"] = "=SUM(Sales!D2:D10)"
        wb.save(str(data_workbook))
        wb.close()

        # Export Sales sheet
        sales_csv = tmp_path / "sales.csv"
        run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(sales_csv),
            "--sheet",
            "Sales",
        )

        # Export Summary sheet
        summary_csv = tmp_path / "summary.csv"
        run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(summary_csv),
            "--sheet",
            "Summary",
        )

        # Verify different exports
        with open(sales_csv) as f:
            sales_content = f.read()
        with open(summary_csv) as f:
            summary_content = f.read()

        assert "Product" in sales_content
        assert "Total Sales" in summary_content


class TestExportPerformance:
    """Test export performance with larger datasets."""

    @pytest.mark.slow
    def test_large_csv_export(self, tmp_path: Path):
        """Test CSV export with many rows."""
        # Create larger workbook
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws["A1"] = "ID"
        ws["B1"] = "Value"
        for i in range(2, 1002):  # 1000 rows
            ws[f"A{i}"] = i - 1
            ws[f"B{i}"] = f"Value {i - 1}"

        path = tmp_path / "large.xlsx"
        wb.save(str(path))

        # Export to CSV
        csv_path = tmp_path / "large.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(path),
            "--outfile",
            str(csv_path),
        )

        assert exit_code == 0
        # Row count includes header + 1000 data rows = 1001 total
        assert output["data"]["row_count"] == 1001

        # Verify CSV size
        assert csv_path.stat().st_size > 0

    @pytest.mark.slow
    def test_large_json_export(self, tmp_path: Path):
        """Test JSON export with many rows."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws["A1"] = "ID"
        ws["B1"] = "Value"
        for i in range(2, 1002):
            ws[f"A{i}"] = i - 1
            ws[f"B{i}"] = f"Value {i - 1}"

        path = tmp_path / "large.xlsx"
        wb.save(str(path))

        json_path = tmp_path / "large.json"
        output, exit_code = run_tool(
            "xls_export_json",
            "--input",
            str(path),
            "--outfile",
            str(json_path),
            "--orient",
            "records",
        )

        assert exit_code == 0
        assert output["data"]["record_count"] == 1000


class TestExportWithFormatting:
    """Test export with formatted data."""

    def test_csv_preserves_values_not_formatting(self, tmp_path: Path):
        """Test that CSV export preserves values but loses formatting."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws["A1"] = "Formatted Number"
        ws["A2"] = 1234.5678

        # Apply formatting
        # NumberFormat removed - not needed

        ws["A2"].number_format = "#,##0.00"

        path = tmp_path / "formatted.xlsx"
        wb.save(str(path))

        # Export to CSV
        csv_path = tmp_path / "output.csv"
        run_tool(
            "xls_export_csv",
            "--input",
            str(path),
            "--outfile",
            str(csv_path),
        )

        with open(csv_path) as f:
            content = f.read()

        # CSV has value, not formatted string
        assert "1234.5678" in content or "1234.57" in content

    def test_json_date_formatting(self, tmp_path: Path):
        """Test that dates are properly formatted in JSON."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws["A1"] = "Date"
        ws["A2"] = datetime(2024, 12, 25, 14, 30, 0)

        path = tmp_path / "dates.xlsx"
        wb.save(str(path))

        json_path = tmp_path / "output.json"
        run_tool(
            "xls_export_json",
            "--input",
            str(path),
            "--outfile",
            str(json_path),
        )

        with open(json_path) as f:
            data = json.load(f)

        # Dates should be ISO formatted
        date_str = str(data[0]["Date"])
        assert "2024-12-25" in date_str or "2024-12-25T14:30:00" in date_str


class TestExportErrorHandling:
    """Test export error handling."""

    def test_corrupted_file_handling(self, tmp_path: Path):
        """Test graceful handling of corrupted files."""
        # Create fake corrupted file
        corrupted = tmp_path / "corrupted.xlsx"
        corrupted.write_bytes(b"This is not a valid Excel file")

        output_path = tmp_path / "output.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(corrupted),
            "--outfile",
            str(output_path),
        )

        # Should fail gracefully
        assert exit_code != 0

    def test_invalid_sheet_name(self, data_workbook: Path, tmp_path: Path):
        """Test error on invalid sheet name."""
        output_path = tmp_path / "output.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
            "--sheet",
            "NonExistentSheet",
        )

        assert exit_code == 1
        assert "sheet" in str(output.get("warnings", [])).lower()

    def test_permission_error(self, data_workbook: Path, tmp_path: Path):
        """Test error on permission denied."""
        import os

        # Root can create any directory, so this test is only meaningful for non-root
        if os.getuid() == 0:
            pytest.skip("Root bypasses permission checks — test requires non-root user")

        output_path = Path("/nonexistent_dir/output.csv")
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
        )

        # Should fail (exit code depends on error type)
        assert exit_code != 0
